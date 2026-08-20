import streamlit as st
import pandas as pd
from sqlalchemy import create_engine, text
from datetime import datetime, timedelta, date, timezone
import calendar
import io
import json
import zipfile
from weasyprint import HTML
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- 1. Database Connection & Auto-Migration ---
try:
    if "DB_URL" not in st.secrets:
        st.error("❌ `DB_URL` missing in Streamlit Secrets! Please configure it in Settings -> Secrets.")
        st.stop()
    
    DB_URL = st.secrets["DB_URL"]
    engine = create_engine(DB_URL)
    
    with engine.begin() as conn:
        try:
            existing_cols = pd.read_sql("SELECT * FROM companies LIMIT 0", conn).columns
            if 'biz_name' not in existing_cols: conn.execute(text("ALTER TABLE companies ADD COLUMN biz_name TEXT"))
            if 'branch_code' not in existing_cols: conn.execute(text("ALTER TABLE companies ADD COLUMN branch_code TEXT"))
            if 'ar_ref_date' not in existing_cols: conn.execute(text("ALTER TABLE companies ADD COLUMN ar_ref_date DATE"))
            if 'br_ref_date' not in existing_cols: conn.execute(text("ALTER TABLE companies ADD COLUMN br_ref_date DATE"))
            if 'cessation_date' not in existing_cols: conn.execute(text("ALTER TABLE companies ADD COLUMN cessation_date DATE"))
            if 'agent' not in existing_cols: conn.execute(text("ALTER TABLE companies ADD COLUMN agent TEXT"))
            if 'year_end' not in existing_cols: conn.execute(text("ALTER TABLE companies ADD COLUMN year_end TEXT"))
            if 'billing_mode' not in existing_cols: conn.execute(text("ALTER TABLE companies ADD COLUMN billing_mode TEXT"))
        except Exception:
            pass

except Exception as db_err:
    st.error("### 🛑 Database Connection Critical Failure")
    st.info(f"**Error Details:**\n`{str(db_err)}`")
    st.stop()

# --- 2. Utility Functions & Dynamic Years (HKT Timezone locked) ---
HKT = timezone(timedelta(hours=8))

def to_date(val):
    try:
        if pd.isna(val) or val == "" or str(val).strip() == "" or str(val).lower() in ["none", "nat", "nan"]: return None
        return pd.to_datetime(val).date()
    except: return None

def clean_val(v):
    v = str(v).strip()
    if v.lower() in ["nat", "none", "nan", ""]: return ""
    if v.endswith(" 00:00:00"): return v.replace(" 00:00:00", "")
    return v.strip()

def get_anniv(year, month, day):
    try: return date(year, month, day)
    except ValueError: return date(year, month, day - 1)

def clean_status(val):
    v = str(val)
    for emo in ["🔴 ", "🟡 ", "🟢 ", "✅ ", "⚪ ", "⏳ ", "⚠️ ", "🆕 ", "💰 ", "📄 ", "✨", "✨ "]: 
        v = v.replace(emo, "")
    return v.strip()

def add_months(dt, months):
    if not dt: return None
    m = dt.month + months
    y = dt.year + (m - 1) // 12
    m = (m - 1) % 12 + 1
    d = min(dt.day, calendar.monthrange(y, m)[1])
    return date(y, m, d)

def get_base_date(row_dict):
    place = str(row_dict.get('incorp_place', ''))
    is_hk_reg = str(row_dict.get('is_hk_registered', 'False')).strip().lower() in ['true', 'yes', 'y', '1']
    if place == 'HK': return to_date(row_dict.get('incorp_date'))
    if is_hk_reg: return to_date(row_dict.get('hk_incorp_date'))
    return None

def calc_ar_deadline(base_dt, ar_ref_date_val, year):
    if ar_ref_date_val:
        base_for_calc = get_anniv(year, ar_ref_date_val.month, ar_ref_date_val.day)
    else:
        base_for_calc = get_anniv(year, base_dt.month, base_dt.day)
    return base_for_calc + timedelta(days=42)

def calc_bvi_fee_deadline(incorp_dt, year):
    if not incorp_dt: return date(year, 5, 31)
    if incorp_dt.month <= 6: return date(year, 5, 31)
    else: return date(year, 11, 30)

def calc_afr_deadline(year_end_str, year):
    if not year_end_str or str(year_end_str).strip() == "": year_end_str = "12/31"
    try: m, d = map(int, str(year_end_str).split('/'))
    except: m, d = 12, 31
    ye_date = get_anniv(year, m, d)
    return add_months(ye_date, 9)

def calc_es_deadline(incorp_dt, year):
    if not incorp_dt: return None
    if incorp_dt.year < 2019:
        return date(year, 12, 31)
    else:
        fp_end = get_anniv(year, incorp_dt.month, incorp_dt.day)
        return add_months(fp_end, 6)

current_system_year = datetime.now(HKT).year
active_years = list(range(2025, current_system_year + 5))
report_years = [y for y in active_years if y <= current_system_year]

# --- 3. Navigation ---
st.set_page_config(page_title="Secretary ERP - V221", layout="wide")
choice = st.sidebar.radio("Navigation (V221)", ["📊 Dashboard", "🏢 Company Register", "⚙️ Group Management", "📤 Data Exchange"])

TEMPLATE_COLS = [
    "client_group", "name_en", "name_ch", "biz_name", "incorp_place", "incorp_place_others", 
    "incorp_date", "ci_no", "is_hk_registered", "hk_incorp_date", "hk_ci_no", "br_no", "branch_code",
    "co_type", "reg_addr", "corres_addr", "round_loc", "sign_loc", "seal_loc", 
    "br_ref_date", "ar_ref_date", "cessation_date", "agent", "year_end", "billing_mode",
    "nd2a_eff_date", "nd2a_file_date", "nd2a_download", "nd4_eff_date", "nd4_file_date", "nd4_download", 
    "nn6_eff_date", "nn6_file_date", "nn6_download",
    "dissolution_date", "remark"
]

EXCHANGE_COL_MAPPING = {
    'client_group': 'Client Group', 'name_en': 'Company Name EN', 'name_ch': 'Company Name CH', 
    'biz_name': 'Business Name', 'incorp_place': 'Incorp Place', 'incorp_place_others': 'Incorp Place Others', 
    'incorp_date': 'Incorp Date', 'ci_no': 'CI No.', 'is_hk_registered': 'Non-HK Registered in HK', 
    'hk_incorp_date': 'HK Incorp Date', 'hk_ci_no': 'HK CI No.', 'br_no': 'BR No.', 
    'branch_code': 'Branch Code', 'co_type': 'Company Type', 'reg_addr': 'Registered Address', 
    'corres_addr': 'Correspondence Address', 'round_loc': 'Round Stamp', 'sign_loc': 'Signature Chop', 
    'seal_loc': 'Common Seal', 'br_ref_date': 'BR Ref Date', 'ar_ref_date': 'AR Ref Date', 
    'cessation_date': 'Cessation Date', 'agent': 'Registered Agent', 'year_end': 'Financial Year End', 'billing_mode': 'Billing Mode',
    'nd2a_eff_date': 'ND2A Eff Date', 'nd2a_file_date': 'ND2A File Date', 
    'nd2a_download': 'ND2A Download', 'nd4_eff_date': 'ND4 Eff Date', 'nd4_file_date': 'ND4 File Date', 
    'nd4_download': 'ND4 Download', 'nn6_eff_date': 'NN6 Eff Date', 'nn6_file_date': 'NN6 File Date', 
    'nn6_download': 'NN6 Download', 'dissolution_date': 'Dissolution Date', 'remark': 'Remark'
}
REVERSE_EXCHANGE_MAPPING = {v: k for k, v in EXCHANGE_COL_MAPPING.items()}
REVERSE_EXCHANGE_MAPPING['Business Name (業務名稱)'] = 'biz_name'
REVERSE_EXCHANGE_MAPPING['BR No. (8-digit)'] = 'br_no'

# --- 4. Report Generation ---

@st.cache_data(show_spinner=False)
def generate_custom_pdf(selected_df, hide_client_group=False):
    now = datetime.now(HKT).strftime("%Y/%m/%d %H:%M")
    def fmt_date(val):
        d = to_date(val)
        return d.strftime('%Y/%m/%d') if d else "N/A"
    
    if selected_df.empty: return b""
    
    sort_cols = [c for c in ['client_group', 'name_en', 'branch_code', 'incorp_place'] if c in selected_df.columns]
    selected_df = selected_df.sort_values(by=sort_cols, na_position='last')

    html_head = f"""
    <html>
    <head>
        <meta charset="UTF-8">
        <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+TC:wght@400;700&display=swap" rel="stylesheet">
        <style>
            @page {{ 
                size: A4; margin: 15mm 10mm 20mm 10mm; 
                @bottom-left {{
                    content: "Company Report | Generated on: {now}";
                    font-size: 8.5pt; color: #7f8c8d; font-family: 'Noto Sans TC', sans-serif;
                }}
                @bottom-right {{
                    content: counter(page) " of " counter(pages) " Page(s)";
                    font-size: 8.5pt; color: #7f8c8d; font-family: 'Noto Sans TC', sans-serif;
                }}
            }}
            body {{ font-family: 'Noto Sans TC', sans-serif; color: #2c3e50; line-height: 1.4; background-color: #ffffff; text-align: justify; }}
            .company-container {{ width: 100%; }}
            .main-table {{ width: 100%; border-collapse: collapse; }}
            .header-content {{ text-align: center; border-bottom: 1px solid #eee; padding-bottom: 15px; margin-bottom: 20px; }}
            .name-en {{ font-size: 20pt; font-weight: bold; color: #2980b9; text-align: center; }}
            .name-ch {{ font-size: 15pt; color: #333333; margin-top: 5px; text-align: center; min-height: 20px; }}
            .section-bar {{ background-color: #f1f4f6; padding: 8px 15px; font-weight: bold; font-size: 11pt; margin: 20px 0 10px 0; border-left: 5px solid #3498db; color: #2c3e50; text-align: left; }}
            .section-group {{ page-break-inside: avoid; }}
            .info-table {{ width: 100%; border-collapse: collapse; }}
            .info-table tr {{ border-bottom: 1px solid #f1f2f6; page-break-inside: avoid; }}
            .info-table th {{ text-align: left; width: 45%; color: #7f8c8d; padding: 8px 0; font-weight: normal; font-size: 10.5pt; }}
            .info-table td {{ text-align: justify; padding: 8px 0; color: #2c3e50; font-size: 10.5pt; font-weight: bold; }}
        </style>
    </head>
    <body>
    """

    cg_row_html = "" if hide_client_group else "<tr><th>Client Group</th><td>__CLIENT_GROUP__</td></tr>"
    docs = []
    
    companies_with_branches = set(selected_df[selected_df['branch_code'] != '000']['name_en'])
    
    for _, row in selected_df.iterrows():
        ch_name = row.get('name_ch', '')
        if not ch_name or pd.isna(ch_name): ch_name = ''
        
        biz_name = str(row.get('biz_name', '')).strip()
        biz_html = f'<div class="name-ch" style="font-size: 11.5pt; color: #7f8c8d; font-weight: normal; margin-top: 4px;">Business Name: {biz_name}</div>' if biz_name and biz_name not in ['None', 'nan'] else ''
        
        place = str(row.get('incorp_place', ''))
        is_hk_reg = str(row.get('is_hk_registered', 'False')).strip().lower() in ['true', 'yes', '1']
        is_bvi = place not in ['HK', ''] and not is_hk_reg
        base_date = get_base_date(row)
        if not base_date and place != 'HK': base_date = to_date(row.get('incorp_date'))
        incorp_year = base_date.year if base_date else None
        
        branch = str(row.get('branch_code', '000')).strip()
        if branch in ['None', 'nan', '', '<NA>']: branch = '000'
        is_branch = branch != '000'
        has_branch = row.get('name_en') in companies_with_branches
        
        disp_en = str(row.get('name_en', ''))
            
        dynamic_place_rows = ""
        display_place = place
        if place == 'Others': display_place = f"Others ({str(row.get('incorp_place_others', ''))})"
            
        dynamic_place_rows += f"<tr><th>{place} Incorp Date</th><td>{fmt_date(row.get('incorp_date'))}</td></tr>"
        dynamic_place_rows += f"<tr><th>{place} CI No.</th><td>{str(row.get('ci_no', ''))}</td></tr>"

        if is_bvi:
            agent = str(row.get('agent', ''))
            ye = str(row.get('year_end', '12/31'))
            bm = str(row.get('billing_mode', ''))
            if 'All-in' in bm: bm = 'All-in Package'
            elif 'Itemized' in bm: bm = 'Itemized'
            
            if agent and agent != 'None': dynamic_place_rows += f"<tr><th>Registered Agent</th><td>{agent}</td></tr>"
            if ye and ye != 'None': dynamic_place_rows += f"<tr><th>Financial Year End</th><td>{ye}</td></tr>"
            if bm and bm != 'None': dynamic_place_rows += f"<tr><th>Billing Mode</th><td>{bm}</td></tr>"

        dynamic_hk_rows = ""
        dynamic_annual_rows = ""
        
        br_ref_raw = to_date(row.get('br_ref_date'))
        ar_ref_raw = to_date(row.get('ar_ref_date'))
        
        if br_ref_raw and (not base_date or (br_ref_raw.month != base_date.month or br_ref_raw.day != base_date.day)):
            dynamic_place_rows += f"<tr><th>BR Ref Date (MM/DD)</th><td>{br_ref_raw.strftime('%m/%d')}</td></tr>"
        if ar_ref_raw and (not base_date or (ar_ref_raw.month != base_date.month or ar_ref_raw.day != base_date.day)):
            dynamic_place_rows += f"<tr><th>AR Ref Date (MM/DD)</th><td>{ar_ref_raw.strftime('%m/%d')}</td></tr>"
            
        br_no_raw = str(row.get('br_no', '')).strip()
        if has_branch and br_no_raw:
            disp_br = f"{br_no_raw}-{branch}"
        else:
            disp_br = br_no_raw
            
        if place == 'HK' or is_hk_reg:
            if place == 'HK':
                dynamic_hk_rows += f"<tr><th>HK BR No.</th><td>{disp_br}</td></tr>"
            else:
                dynamic_hk_rows += f"<tr><th>HK Incorp Date</th><td>{fmt_date(row.get('hk_incorp_date'))}</td></tr>"
                dynamic_hk_rows += f"<tr><th>HK CI No.</th><td>{str(row.get('hk_ci_no', ''))}</td></tr>"
                dynamic_hk_rows += f"<tr><th>HK BR No.</th><td>{disp_br}</td></tr>"

        comp_rec_str = str(row.get('compliance_records', '{}'))
        try: rec_dict = json.loads(comp_rec_str)
        except: rec_dict = {}
        if not isinstance(rec_dict, dict): rec_dict = {}
        
        cess_date = to_date(row.get('cessation_date'))
        year_display_data = {}
        
        prev_br_by = 'Firm'
        prev_afr_fee_by = 'Firm'
        prev_es_fee_by = 'Firm'
        
        for y in report_years:
            y_str = str(y)
            y_data = rec_dict.get(y_str, {})
            
            if incorp_year and y < incorp_year:
                year_display_data[y] = {'not_incorp': True}
                br_by = 'N/A'
                afr_fee_by = 'N/A'
                es_fee_by = 'N/A'
            else:
                raw_br_by = str(y_data.get('fee_by', y_data.get('br_paid_by', ''))).strip()
                if raw_br_by: br_by = raw_br_by
                else: br_by = prev_br_by if prev_br_by != 'N/A' else 'Firm'
                
                raw_afr_fee_by = str(y_data.get('afr_fee_by', '')).strip()
                if raw_afr_fee_by: afr_fee_by = raw_afr_fee_by
                else: afr_fee_by = prev_afr_fee_by if prev_afr_fee_by != 'N/A' else 'Firm'
                
                raw_es_fee_by = str(y_data.get('es_fee_by', '')).strip()
                if raw_es_fee_by: es_fee_by = raw_es_fee_by
                else: es_fee_by = prev_es_fee_by if prev_es_fee_by != 'N/A' else 'Firm'
                
                if is_branch and cess_date and y >= cess_date.year:
                    br_by = "N/A"
                
                br_dt = y_data.get('fee_date', y_data.get('br_date', 'N/A'))
                if br_dt is None or str(br_dt).strip() in ['None', '']: br_dt = 'N/A'
                else: br_dt = str(br_dt)
                
                afr_fee_dt = y_data.get('afr_fee_date', 'N/A')
                if afr_fee_dt is None or str(afr_fee_dt).strip() in ['None', '']: afr_fee_dt = 'N/A'
                else: afr_fee_dt = str(afr_fee_dt)
                
                ar_dt = y_data.get('ar_date', 'N/A')
                if ar_dt is None or str(ar_dt).strip() in ['None', '']: ar_dt = 'N/A'
                else: ar_dt = str(ar_dt)
                
                ar_cr_status = y_data.get('ar_cr_status', 'Pending')
                
                es_fee_dt = y_data.get('es_fee_date', 'N/A')
                if es_fee_dt is None or str(es_fee_dt).strip() in ['None', '']: es_fee_dt = 'N/A'
                else: es_fee_dt = str(es_fee_dt)
                
                es_dt = y_data.get('es_date', 'N/A')
                if es_dt is None or str(es_dt).strip() in ['None', '']: es_dt = 'N/A'
                else: es_dt = str(es_dt)
                
                if ar_cr_status == 'Exempt (Dormant)':
                    ar_dt_disp = "N/A"
                    ar_cr_disp = "Exempt (Dormant)"
                elif ar_cr_status == 'Included in Agent Fee':
                    ar_dt_disp = "N/A"
                    ar_cr_disp = "Included in Agent Fee"
                elif is_branch:
                    ar_dt_disp = "N/A (Branch)"
                    ar_cr_disp = "N/A"
                elif incorp_year and y == incorp_year and ar_dt == 'N/A':
                    ar_dt_disp = "Exempt (1st Year)"
                    ar_cr_disp = "N/A"
                else:
                    ar_dt_disp = ar_dt
                    ar_cr_disp = ar_cr_status
                    
                if incorp_year and y == incorp_year and es_dt == 'N/A':
                    es_dt_disp = "Exempt (1st Year)"
                else:
                    es_dt_disp = es_dt
                    
                if is_branch and cess_date and y >= cess_date.year:
                    br_dt = "N/A"
                    
                year_display_data[y] = {
                    'not_incorp': False,
                    'br_by': br_by,
                    'br_dt': br_dt,
                    'afr_fee_by': afr_fee_by,
                    'afr_fee_dt': afr_fee_dt,
                    'ar_dt_disp': ar_dt_disp,
                    'ar_cr_disp': ar_cr_disp,
                    'es_fee_by': es_fee_by,
                    'es_fee_dt': es_fee_dt,
                    'es_dt_disp': es_dt_disp
                }
                
            prev_br_by = br_by
            prev_afr_fee_by = afr_fee_by
            prev_es_fee_by = es_fee_by

        for y in reversed(report_years):
            yd = year_display_data[y]
            
            if y == current_system_year:
                header_html = f"<tr><th colspan='2' style='background-color:#ebf5fb; color:#2980b9; text-align:center; border: 1px solid #aed6f1; padding: 6px 0; font-weight: bold;'>=== Year {y} Annual Cycle (Current) ===</th></tr>"
            else:
                header_html = f"<tr><th colspan='2' style='background-color:#fcfcfc; color:#95a5a6; text-align:center; font-weight:normal; border-top: 1px dashed #eaeded; padding: 6px 0;'>--- Year {y} Annual Cycle ---</th></tr>"
            
            dynamic_annual_rows += header_html
            
            if yd['not_incorp']:
                dynamic_annual_rows += f"<tr><td colspan='2' style='text-align:center; color:#bdc3c7; font-weight:normal;'>Not Incorporated Yet</td></tr>"
            else:
                br_dt_disp = yd['br_dt']
                if br_dt_disp != 'N/A': br_dt_disp = br_dt_disp.replace('-', '/')
                
                afr_fee_dt_disp = yd['afr_fee_dt']
                if afr_fee_dt_disp != 'N/A': afr_fee_dt_disp = afr_fee_dt_disp.replace('-', '/')
                
                ar_dt_disp = yd['ar_dt_disp']
                if ar_dt_disp not in ['N/A', 'Exempt (1st Year)', 'N/A (Branch)']: ar_dt_disp = ar_dt_disp.replace('-', '/')
                
                es_fee_dt_disp = yd['es_fee_dt']
                if es_fee_dt_disp != 'N/A': es_fee_dt_disp = es_fee_dt_disp.replace('-', '/')
                    
                es_dt_disp = yd['es_dt_disp']
                if es_dt_disp not in ['N/A', 'Exempt (1st Year)']: es_dt_disp = es_dt_disp.replace('-', '/')
                    
                text_color = "#2c3e50" if y == current_system_year else "#5d6d7e"
                font_weight = "bold" if y == current_system_year else "normal"
                
                if is_branch and cess_date and y >= cess_date.year:
                    dynamic_annual_rows += f"<tr><th colspan='2' style='text-align:center; color:#e74c3c; font-weight:bold; background-color:#fdedec;'>N/A (Branch Cessed)</th></tr>"
                else:
                    lbl_br = "Annual Fee Paid By" if is_bvi else "BR Paid By"
                    lbl_br_dt = "Annual Fee Paid Date" if is_bvi else "BR Paid Date"
                    lbl_ar_dt = "AFR Filed Date" if is_bvi else "AR Filed Date"
                    lbl_ar_st = "AFR Status" if is_bvi else "AR CR Status"
                    
                    dynamic_annual_rows += f"<tr><th>{lbl_br} ({y})</th><td style='color:{text_color}; font-weight:{font_weight};'>{yd['br_by']}</td></tr>"
                    dynamic_annual_rows += f"<tr><th>{lbl_br_dt} ({y})</th><td style='color:{text_color}; font-weight:{font_weight};'>{br_dt_disp}</td></tr>"
                    if is_bvi:
                        dynamic_annual_rows += f"<tr><th>AFR Fee Paid By ({y})</th><td style='color:{text_color}; font-weight:{font_weight};'>{yd['afr_fee_by']}</td></tr>"
                        dynamic_annual_rows += f"<tr><th>AFR Fee Paid Date ({y})</th><td style='color:{text_color}; font-weight:{font_weight};'>{afr_fee_dt_disp}</td></tr>"
                    dynamic_annual_rows += f"<tr><th>{lbl_ar_dt} ({y})</th><td style='color:{text_color}; font-weight:{font_weight};'>{ar_dt_disp}</td></tr>"
                    if not is_bvi:
                        dynamic_annual_rows += f"<tr><th>{lbl_ar_st} ({y})</th><td style='color:{text_color}; font-weight:{font_weight};'>{yd['ar_cr_disp']}</td></tr>"
                    else:
                        dynamic_annual_rows += f"<tr><th>ES Fee Paid By ({y})</th><td style='color:{text_color}; font-weight:{font_weight};'>{yd['es_fee_by']}</td></tr>"
                        dynamic_annual_rows += f"<tr><th>ES Fee Paid Date ({y})</th><td style='color:{text_color}; font-weight:{font_weight};'>{es_fee_dt_disp}</td></tr>"
                        dynamic_annual_rows += f"<tr><th>ES Filed Date ({y})</th><td style='color:{text_color}; font-weight:{font_weight};'>{es_dt_disp}</td></tr>"
        
        n2e_val = to_date(row.get('nd2a_eff_date'))
        n2f_val = to_date(row.get('nd2a_file_date'))
        n4e_val = to_date(row.get('nd4_eff_date'))
        n4f_val = to_date(row.get('nd4_file_date'))
        nn6_e_val = to_date(row.get('nn6_eff_date'))
        nn6_f_val = to_date(row.get('nn6_file_date'))

        has_hk_sec = bool(n2e_val or n2f_val or n4e_val or n4f_val)
        has_nonhk_sec = bool(nn6_e_val or nn6_f_val)

        dynamic_sec_rows = ""
        if place == 'HK' and not is_branch:
            if has_hk_sec:
                dynamic_sec_rows += f"""
                <div class="section-group">
                    <div class="section-bar">Company Secretary Actions</div>
                    <table class="info-table">
                        <tr><th>ND2A Eff Date</th><td>{fmt_date(row.get('nd2a_eff_date'))}</td></tr>
                        <tr><th>ND4 Eff Date</th><td>{fmt_date(row.get('nd4_eff_date'))}</td></tr>
                    </table>
                </div>"""
        elif is_hk_reg and not is_branch:
            if has_nonhk_sec:
                dynamic_sec_rows += f"""
                <div class="section-group">
                    <div class="section-bar">Non-HK Company Secretary Actions</div>
                    <table class="info-table">
                        <tr><th>NN6 Eff Date</th><td>{fmt_date(row.get('nn6_eff_date'))}</td></tr>
                    </table>
                </div>"""

        remark_val = str(row.get('remark', ''))
        if remark_val == 'None' or not remark_val: remark_val = 'No remarks.'

        this_cg = "" if hide_client_group else cg_row_html.replace('__CLIENT_GROUP__', str(row.get('client_group', '')))
        
        card_html = f"""
        <div class="company-container">
            <table class="main-table">
                <thead>
                    <tr><td><div class="header-content"><div class="name-en">{disp_en}</div><div class="name-ch">{str(ch_name)}</div>{biz_html}</div></td></tr>
                </thead>
                <tbody>
                    <tr>
                        <td>
                            <div class="company-card">
                                <div class="section-group">
                                    <div class="section-bar">Registration Details</div>
                                    <table class="info-table">
                                        {this_cg}
                                        <tr><th>Incorp Place</th><td>{display_place}</td></tr>
                                        {dynamic_place_rows}
                                        {dynamic_hk_rows}
                                        <tr><th>Company Type</th><td>{str(row.get('co_type', ''))}</td></tr>
                                    </table>
                                </div>
                                <div class="section-group">
                                    <div class="section-bar">Addresses</div>
                                    <table class="info-table">
                                        <tr><th>Registered Address</th><td>{str(row.get('reg_addr', ''))}</td></tr>
                                        <tr><th>Correspondence Address</th><td>{str(row.get('corres_addr', ''))}</td></tr>
                                    </table>
                                </div>
                                <div class="section-group">
                                    <div class="section-bar">Items Storage</div>
                                    <table class="info-table">
                                        <tr><th>Round Stamp</th><td>{str(row.get('round_loc', ''))}</td></tr>
                                        <tr><th>Signature Chop</th><td>{str(row.get('sign_loc', ''))}</td></tr>
                                        <tr><th>Common Seal</th><td>{str(row.get('seal_loc', ''))}</td></tr>
                                    </table>
                                </div>
                                
                                <div class="section-group">
                                    <div class="section-bar">Remarks</div>
                                    <div style="padding: 8px 15px; font-size: 10.5pt; color: #2c3e50; white-space: pre-wrap;">{remark_val}</div>
                                </div>
                                
                                {dynamic_sec_rows}
                                
                                <div class="section-group">
                                    <div class="section-bar">Compliance Filings (Yearly)</div>
                                    <table class="info-table">
                                        {dynamic_annual_rows}
                                    </table>
                                </div>
                            </div>
                        </td>
                    </tr>
                </tbody>
            </table>
        </div>
        """
        full_doc = html_head + card_html + "</body></html>"
        docs.append(HTML(string=full_doc).render())

    if docs:
        all_pages = []
        for doc in docs:
            all_pages.extend(doc.pages)
        buf = io.BytesIO()
        docs[0].copy(all_pages).write_pdf(buf)
        return buf.getvalue()
    return b""

@st.cache_data(show_spinner=False)
def generate_general_excel(selected_df, hide_client_group=False):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Arial", size=10)
    font_red = Font(name="Arial", size=10, color="FF0000", bold=True)
    font_yellow = Font(name="Arial", size=10, color="FF9900", bold=True)
    font_green = Font(name="Arial", size=10, color="00B050", bold=True)
    font_grey = Font(name="Arial", size=10, color="7F8C8D", italic=True)

    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    groups = selected_df['client_group'].unique() if 'client_group' in selected_df.columns else ['Companies']
    groups = sorted([g for g in groups if pd.notna(g)])
    if not groups: groups = ['Companies']

    for g in groups:
        safe_title = str(g).replace("/", "_").replace("\\", "_").replace("[", "").replace("]", "").replace("*", "").replace(":", "").replace("?", "")
        safe_title = safe_title[:31] if safe_title.strip() else "Ungrouped"
        ws = wb.create_sheet(title=safe_title)
        ws.views.sheetView[0].showGridLines = True
        
        group_df = selected_df[selected_df['client_group'] == g] if 'client_group' in selected_df.columns else selected_df
        export_records = group_df.to_dict('records')
        processed_records = []
        
        for row in export_records:
            base_date = get_base_date(row)
            if not base_date and row.get('incorp_place') not in ['HK', '']: base_date = to_date(row.get('incorp_date'))
            incorp_year = base_date.year if base_date else None
            
            branch = str(row.get('branch_code', '000')).strip()
            if branch in ['None', 'nan', '', '<NA>']: branch = '000'
            is_branch = branch != '000'
            cess_date = to_date(row.get('cessation_date'))
            
            row['Branch Code'] = branch 
            row['Business Name'] = row.get('biz_name', '')
            
            br_no_raw = str(row.get('br_no', '')).strip()
            if br_no_raw:
                companies_with_branches = set(selected_df[selected_df['branch_code'] != '000']['name_en'])
                if row.get('name_en') in companies_with_branches:
                    row['BR No.'] = f"{br_no_raw}-{branch}"
                else:
                    row['BR No.'] = br_no_raw
            else:
                 row['BR No.'] = ''
                 
            comp_rec_str = str(row.get('compliance_records', '{}'))
            try: rec_dict = json.loads(comp_rec_str)
            except: rec_dict = {}
            if not isinstance(rec_dict, dict): rec_dict = {}
            
            prev_br_by = 'Firm'
            prev_afr_fee_by = 'Firm'
            prev_es_fee_by = 'Firm'
            
            for y in report_years:
                if incorp_year and y < incorp_year:
                    row[f'{y} Fee Paid By'] = 'N/A'
                    row[f'{y} AR/AFR Status'] = 'N/A'
                    row[f'{y} Fee Paid Date'] = ''
                    row[f'{y} AR/AFR Fee Paid By'] = 'N/A'
                    row[f'{y} AR/AFR Fee Paid Date'] = ''
                    row[f'{y} AR/AFR Filed Date'] = ''
                    row[f'{y} ES Fee Paid By'] = 'N/A'
                    row[f'{y} ES Fee Paid Date'] = ''
                    row[f'{y} ES Filed Date'] = ''
                    row[f'{y} ES Status'] = 'N/A'
                    prev_br_by, prev_afr_fee_by, prev_es_fee_by = 'N/A', 'N/A', 'N/A'
                    continue
                    
                y_str = str(y)
                y_data = rec_dict.get(y_str, {})
                
                raw_br_by = str(y_data.get('fee_by', y_data.get('br_paid_by', ''))).strip()
                if raw_br_by: br_by = raw_br_by
                else: br_by = prev_br_by if prev_br_by != 'N/A' else 'Firm'
                
                raw_afr_fee_by = str(y_data.get('afr_fee_by', '')).strip()
                if raw_afr_fee_by: afr_fee_by = raw_afr_fee_by
                else: afr_fee_by = prev_afr_fee_by if prev_afr_fee_by != 'N/A' else 'Firm'
                
                raw_es_fee_by = str(y_data.get('es_fee_by', '')).strip()
                if raw_es_fee_by: es_fee_by = raw_es_fee_by
                else: es_fee_by = prev_es_fee_by if prev_es_fee_by != 'N/A' else 'Firm'
                
                if is_branch and cess_date and y >= cess_date.year:
                    br_by = "N/A"
                    
                prev_br_by = br_by
                prev_afr_fee_by = afr_fee_by
                prev_es_fee_by = es_fee_by
                
                row[f'{y} Fee Paid By'] = br_by
                
                ar_dt = str(y_data.get('ar_date', ''))
                if ar_dt in ['None', 'nan', '<NA>']: ar_dt = ''
                
                ar_cr_status = y_data.get('ar_cr_status', '')
                if not ar_cr_status:
                    if ar_dt: ar_cr_status = 'Completed'
                    else: ar_cr_status = 'Pending'
                
                es_dt = str(y_data.get('es_date', ''))
                if es_dt in ['None', 'nan', '<NA>']: es_dt = ''
                
                es_status = y_data.get('es_status', '')
                if not es_status:
                    if es_dt: es_status = 'Completed'
                    else: es_status = 'Pending'
                    
                if is_branch:
                    row[f'{y} AR/AFR Status'] = "N/A (Branch)"
                    row[f'{y} ES Status'] = "N/A (Branch)"
                else:
                    row[f'{y} AR/AFR Status'] = ar_cr_status
                    row[f'{y} ES Status'] = es_status
                
                br_d = str(y_data.get('fee_date', y_data.get('br_date', '')))
                if br_d in ['None', 'nan', '<NA>']: br_d = ''
                else: br_d = br_d.replace('-', '/')
                
                afr_fee_dt = str(y_data.get('afr_fee_date', ''))
                if afr_fee_dt in ['None', 'nan', '<NA>']: afr_fee_dt = ''
                else: afr_fee_dt = afr_fee_dt.replace('-', '/')
                
                es_fee_dt = str(y_data.get('es_fee_date', ''))
                if es_fee_dt in ['None', 'nan', '<NA>']: es_fee_dt = ''
                else: es_fee_dt = es_fee_dt.replace('-', '/')
                
                if ar_cr_status in ['Exempt (Dormant)', 'Included in Agent Fee']:
                    ar_dt = ''
                elif ar_dt and not is_branch: ar_dt = ar_dt.replace('-', '/')
                elif is_branch: ar_dt = ''
                
                if es_status == 'Exempt':
                    es_dt = ''
                elif es_dt and not is_branch: es_dt = es_dt.replace('-', '/')
                elif is_branch: es_dt = ''
                
                if is_branch and cess_date and y >= cess_date.year:
                    row[f'{y} Fee Paid Date'] = ""
                else:
                    row[f'{y} Fee Paid Date'] = br_d
                    
                row[f'{y} AR/AFR Fee Paid By'] = afr_fee_by
                row[f'{y} AR/AFR Fee Paid Date'] = afr_fee_dt
                row[f'{y} AR/AFR Filed Date'] = ar_dt
                row[f'{y} ES Fee Paid By'] = es_fee_by
                row[f'{y} ES Fee Paid Date'] = es_fee_dt
                row[f'{y} ES Filed Date'] = es_dt

            for col in ["incorp_date", "hk_incorp_date", "br_ref_date", "ar_ref_date", "cessation_date", "nd2a_eff_date", "nd2a_file_date", "nd4_eff_date", "nd4_file_date", "nn6_eff_date", "nn6_file_date", "dissolution_date"]:
                val = to_date(row.get(col))
                row[col] = val.strftime('%Y/%m/%d') if val else ""

            processed_records.append(row)

        df_export = pd.DataFrame(processed_records)
        
        col_mapping = {
            'client_group': 'Client Group', 'name_en': 'Company Name EN', 'name_ch': 'Company Name CH', 
            'Business Name': 'Business Name', 'incorp_place': 'Incorp Place', 'incorp_place_others': 'Incorp Place Others', 
            'incorp_date': 'Incorp Date', 'ci_no': 'CI No.', 'is_hk_registered': 'Non-HK Registered in HK', 
            'hk_incorp_date': 'HK Incorp Date', 'hk_ci_no': 'HK CI No.', 'BR No.': 'BR No.', 
            'Branch Code': 'Branch Code', 'co_type': 'Company Type', 'reg_addr': 'Registered Address', 
            'corres_addr': 'Correspondence Address', 'round_loc': 'Round Stamp', 'sign_loc': 'Signature Chop', 
            'seal_loc': 'Common Seal', 'br_ref_date': 'BR Ref Date', 'ar_ref_date': 'AR Ref Date', 
            'cessation_date': 'Cessation Date', 'agent': 'Registered Agent', 'year_end': 'Financial Year End', 'billing_mode': 'Billing Mode',
            'nd2a_eff_date': 'ND2A Eff Date', 'nd2a_file_date': 'ND2A File Date', 
            'nd2a_download': 'ND2A Download', 'nd4_eff_date': 'ND4 Eff Date', 'nd4_file_date': 'ND4 File Date', 
            'nd4_download': 'ND4 Download', 'nn6_eff_date': 'NN6 Eff Date', 'nn6_file_date': 'NN6 File Date', 
            'nn6_download': 'NN6 Download', 'dissolution_date': 'Dissolution Date', 'remark': 'Remark'
        }
        
        df_export.rename(columns=col_mapping, inplace=True, errors='ignore')
        
        base_cols_ordered = list(col_mapping.values())
        base_cols_ordered.remove('Remark')
        
        dyn_cols = []
        for y in report_years:
            dyn_cols.extend([f"{y} Fee Paid By", f"{y} Fee Paid Date", f"{y} AR/AFR Fee Paid By", f"{y} AR/AFR Fee Paid Date", f"{y} AR/AFR Filed Date", f"{y} AR/AFR Status", f"{y} ES Fee Paid By", f"{y} ES Fee Paid Date", f"{y} ES Filed Date", f"{y} ES Status"])
        
        headers = [c for c in base_cols_ordered if c in df_export.columns] + [c for c in dyn_cols if c in df_export.columns] + ['Remark']
        if hide_client_group and 'Client Group' in headers:
            headers.remove('Client Group')
            
        ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
        ws["A1"] = f"Company Report - Group: {g} (Generated on: {datetime.now(HKT).strftime('%Y/%m/%d %H:%M')})"
        ws["A1"].font = Font(name="Arial", size=14, bold=True, color="1F497D")
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 30
        
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[3].height = 25
        
        current_row = 4
        for _, item in df_export.iterrows():
            for col_idx, h in enumerate(headers, 1):
                val = str(item.get(h, ""))
                if val in ['nan', 'None', '<NA>']: val = ""
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = font_data
                cell.border = thin_border
                if current_row % 2 == 0: cell.fill = fill_zebra
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                if "Status" in h:
                    if "Overdue" in val or "Returned" in val: cell.font = font_red
                    elif "Due Soon" in val or "Processing" in val: cell.font = font_yellow
                    elif "Exempt" in val or "Pending" in val or "N/A" in val or "Included" in val: cell.font = font_grey
                    elif "Completed" in val: cell.font = font_green

            ws.row_dimensions[current_row].height = 20
            current_row += 1
            
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row == 1: continue 
                if cell.value: max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(min(max_len + 4, 50), 12)

    if len(wb.sheetnames) == 0:
        wb.create_sheet("No Data")
        
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

@st.cache_data(show_spinner=False)
def generate_beautiful_excel(df, hide_client_group=False):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Arial", size=10)
    font_red = Font(name="Arial", size=10, color="FF0000", bold=True)
    font_yellow = Font(name="Arial", size=10, color="FF9900", bold=True)
    font_green = Font(name="Arial", size=10, color="00B050", bold=True)
    font_grey = Font(name="Arial", size=10, color="7F8C8D", italic=True)

    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    headers = ["Client Group", "Company Name EN", "Company Name CH", "Business Name", "Incorp Place", "Year", "Anniversary (MM/DD)", "BR No.", "Fee Paid By", "Fee Paid Date", "Fee Deadline", "Fee Status", "AR/AFR Fee By", "AR/AFR Fee Date", "AR/AFR Filed Date", "AR/AFR Deadline", "AR/AFR Status", "ES Fee By", "ES Fee Date", "ES Filed Date", "ES Deadline", "ES Status", "Remark"]
    if hide_client_group: headers.remove("Client Group")

    groups = df['Client Group'].unique() if 'Client Group' in df.columns else ['Outstanding']
    groups = sorted([g for g in groups if pd.notna(g)])
    if not groups: groups = ['Outstanding']

    for g in groups:
        safe_title = str(g).replace("/", "_").replace("\\", "_").replace("[", "").replace("]", "").replace("*", "").replace(":", "").replace("?", "")
        safe_title = safe_title[:31] if safe_title.strip() else "Ungrouped"
        ws = wb.create_sheet(title=safe_title)
        ws.views.sheetView[0].showGridLines = True

        ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
        ws["A1"] = f"Outstanding Report - Group: {g} (Generated on: {datetime.now(HKT).strftime('%Y/%m/%d %H:%M')})"
        ws["A1"].font = Font(name="Arial", size=14, bold=True, color="1F497D")
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 30

        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[3].height = 25

        center_cols = ["Client Group", "Incorp Place", "Year", "Anniversary (MM/DD)", "BR No.", "Fee Paid By", "Fee Paid Date", "Fee Deadline", "Fee Status", "AR/AFR Fee By", "AR/AFR Fee Date", "AR/AFR Filed Date", "AR/AFR Deadline", "AR/AFR Status", "ES Fee By", "ES Fee Date", "ES Filed Date", "ES Deadline", "ES Status"]

        group_df = df[df['Client Group'] == g] if 'Client Group' in df.columns else df
        
        current_row = 4
        for _, item in group_df.iterrows():
            row_dict = {
                "Client Group": item.get("Client Group", ""),
                "Company Name EN": item.get("Company Name EN", ""),
                "Company Name CH": item.get("Company Name CH", ""),
                "Business Name": item.get("Business Name", ""),
                "Incorp Place": item.get("Incorp Place", ""),
                "Year": str(item.get("Year", "")),
                "Anniversary (MM/DD)": item.get("Anniversary (MM/DD)", ""),
                "BR No.": item.get("BR No.", ""),
                "Fee Paid By": item.get("Fee Paid By", ""),
                "Fee Paid Date": item.get("Fee Paid Date", ""),
                "Fee Deadline": item.get("Fee Deadline", ""),
                "Fee Status": item.get("Fee Status", ""),
                "AR/AFR Fee By": item.get("AR/AFR Fee By", ""),
                "AR/AFR Fee Date": item.get("AR/AFR Fee Date", ""),
                "AR/AFR Filed Date": item.get("AR/AFR Filed Date", ""),
                "AR/AFR Deadline": item.get("AR/AFR Deadline", ""),
                "AR/AFR Status": item.get("AR/AFR Status", ""),
                "ES Fee By": item.get("ES Fee By", ""),
                "ES Fee Date": item.get("ES Fee Date", ""),
                "ES Filed Date": item.get("ES Filed Date", ""),
                "ES Deadline": item.get("ES Deadline", ""),
                "ES Status": item.get("ES Status", ""),
                "Remark": item.get("Remark", "")
            }
            
            for col_idx, h in enumerate(headers, 1):
                val_str = clean_status(row_dict[h]) if row_dict[h] else ""
                cell = ws.cell(row=current_row, column=col_idx, value=val_str)
                cell.font = font_data
                cell.border = thin_border
                if current_row % 2 == 0: cell.fill = fill_zebra
                
                if h in center_cols: cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else: cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    
                if "Status" in h:
                    if "Overdue" in val_str or "Returned" in val_str: cell.font = font_red
                    elif "Due Soon" in val_str or "Processing" in val_str: cell.font = font_yellow
                    elif "Exempt" in val_str or "Not Incorporated" in val_str or "Pending" in val_str or "Branch" in val_str or "Cessed" in val_str or "Included" in val_str: cell.font = font_grey
                    else: cell.font = font_green
                        
            ws.row_dimensions[current_row].height = 20
            current_row += 1

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row == 1: continue 
                if cell.value: max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(min(max_len + 4, 50), 15)

    if len(wb.sheetnames) == 0:
        wb.create_sheet("No Data")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

@st.cache_data(show_spinner=False)
def generate_outstanding_pdf(df, hide_client_group=False):
    now_str = datetime.now(HKT).strftime("%Y/%m/%d %H:%M")
    colspan = "17" if hide_client_group else "18"
    cg_th = "" if hide_client_group else '<th style="width:4%">Client Group</th>'
    
    html = f"""
    <html>
    <head>
        <meta charset="UTF-8">
        <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+TC:wght@400;700&display=swap" rel="stylesheet">
        <style>
            @page {{ 
                size: A4 landscape; margin: 10mm; background-color: #ffffff; 
                @bottom-left {{ content: "Outstanding Report | Generated on: {now_str}"; font-size: 8pt; color: #7f8c8d; font-family: 'Noto Sans TC', sans-serif; }}
                @bottom-right {{ content: counter(page) " of " counter(pages) " Page(s)"; font-size: 8pt; color: #7f8c8d; font-family: 'Noto Sans TC', sans-serif; }}
            }}
            body {{ font-family: 'Noto Sans TC', sans-serif; font-size: 7pt; color: #2c3e50; margin: 0; padding: 0; }}
            table {{ width: 100%; border-collapse: collapse; table-layout: fixed; }}
            thead {{ display: table-header-group; }}
            tr {{ page-break-inside: avoid; }}
            th {{ background-color: #1f497d; color: white; padding: 3px 1px; border: 1px solid #d9d9d9; font-size: 6pt; text-align: center; vertical-align: middle; word-wrap: break-word; }}
            td {{ padding: 3px 1px; border: 1px solid #d9d9d9; text-align: center; vertical-align: middle; word-wrap: break-word; }}
            tr:nth-child(even) td {{ background-color: #f8f9fa; }}
            .text-left {{ text-align: left; font-weight: bold; color: #2980b9; padding-left: 3px; }}
            .report-header-cell {{ border: none !important; background-color: white !important; text-align: left !important; padding: 0 0 10px 0 !important; }}
        </style>
    </head>
    <body>
        <table>
            <thead>
                <tr>
                    <td colspan="{colspan}" class="report-header-cell">
                        <h2 style="color: #1f497d; margin: 0 0 5px 0;">Outstanding Report</h2>
                    </td>
                </tr>
                <tr>
                    {cg_th}
                    <th style="width:10%">Company Name EN</th>
                    <th style="width:3.5%">Place</th>
                    <th style="width:3%">Year</th>
                    <th style="width:4.5%">Anniv<br>(MM/DD)</th>
                    <th style="width:5%">BR No.</th>
                    <th style="width:4%">Fee By</th>
                    <th style="width:5.5%">Fee Date</th>
                    <th style="width:7%">Fee DL & Status</th>
                    <th style="width:4%">AFR Fee By</th>
                    <th style="width:5.5%">AFR Fee Dt</th>
                    <th style="width:5.5%">AFR Filed</th>
                    <th style="width:7%">AFR DL & Status</th>
                    <th style="width:4%">ES Fee By</th>
                    <th style="width:5.5%">ES Fee Dt</th>
                    <th style="width:5.5%">ES Filed</th>
                    <th style="width:7%">ES DL & Status</th>
                    <th>Remark</th>
                </tr>
            </thead>
            <tbody>
    """
    for _, r in df.iterrows():
        br_val = clean_status(r.get('Fee Status', ''))
        ar_val = clean_status(r.get('AR/AFR Status', ''))
        es_val = clean_status(r.get('ES Status', ''))
        
        def get_color(v):
            if "Overdue" in v or "Returned" in v: return "#ff0000"
            if "Due Soon" in v or "Processing" in v: return "#ff9900"
            if "Exempt" in v or "Not Incorporated" in v or "Pending" in v or "Branch" in v or "Cessed" in v or "Included" in v: return "#7f8c8d"
            return "#00b050"
            
        br_color = get_color(br_val)
        ar_color = get_color(ar_val)
        es_color = get_color(es_val)
        cg_td = "" if hide_client_group else f"<td>{r.get('Client Group', '')}</td>"
        
        name_en = r.get('Company Name EN', '')
        biz = str(r.get('Business Name', '')).strip()
        if biz and biz not in ['None', 'nan']:
            name_en += f"<br><span style='font-size: 5pt; color: #7f8c8d; font-weight: normal;'>Business Name: {biz}</span>"
            
        fee_dl_stat = f"{r.get('Fee Deadline', '')}<br><span style='color: {br_color}; font-size: 6pt;'>{br_val}</span>" if r.get('Fee Deadline') not in ["", "N/A"] else r.get('Fee Deadline', '')
        afr_dl_stat = f"{r.get('AR/AFR Deadline', '')}<br><span style='color: {ar_color}; font-size: 6pt;'>{ar_val}</span>" if r.get('AR/AFR Deadline') not in ["", "N/A"] else r.get('AR/AFR Deadline', '')
        es_dl_stat = f"{r.get('ES Deadline', '')}<br><span style='color: {es_color}; font-size: 6pt;'>{es_val}</span>" if r.get('ES Deadline') not in ["", "N/A"] else r.get('ES Deadline', '')
        
        html += f"""
        <tr>
            {cg_td}
            <td class="text-left">{name_en}</td>
            <td>{r.get('Incorp Place', '')}</td>
            <td style="font-weight: bold; color: #1f497d;">{r.get('Year', '')}</td>
            <td style="font-weight: bold;">{r.get('Anniversary (MM/DD)', '')}</td>
            <td style="font-weight: bold;">{r.get('BR No.', '')}</td>
            <td>{r.get('Fee Paid By', '')}</td>
            <td>{r.get('Fee Paid Date', '')}</td>
            <td>{fee_dl_stat}</td>
            <td>{r.get('AR/AFR Fee By', '')}</td>
            <td>{r.get('AR/AFR Fee Date', '')}</td>
            <td>{r.get('AR/AFR Filed Date', '')}</td>
            <td>{afr_dl_stat}</td>
            <td>{r.get('ES Fee By', '')}</td>
            <td>{r.get('ES Fee Date', '')}</td>
            <td>{r.get('ES Filed Date', '')}</td>
            <td>{es_dl_stat}</td>
            <td style="text-align: left; font-size: 6pt; color: #7f8c8d;">{r.get('Remark', '')}</td>
        </tr>
        """
    html += "</tbody></table></body></html>"
    return HTML(string=html).write_pdf()

@st.cache_data(show_spinner=False)
def generate_inv_excel(df, year, month_disp, hide_client_group=False):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Arial", size=10)
    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    headers = ["Client Group", "Company Name EN", "Company Name CH", "Business Name", "Incorp Place", "Year", "Anniversary (MM/DD)", "BR No.", "Fee Paid By", "Billing Item", "Fee Deadline", "AR/AFR Deadline", "ES Deadline", "Remark"]
    if hide_client_group: headers.remove("Client Group")

    groups = df['Client Group'].unique() if 'Client Group' in df.columns else ['Invoicing']
    groups = sorted([g for g in groups if pd.notna(g)])
    if not groups: groups = ['Invoicing']

    for g in groups:
        safe_title = str(g).replace("/", "_").replace("\\", "_").replace("[", "").replace("]", "").replace("*", "").replace(":", "").replace("?", "")
        safe_title = safe_title[:31] if safe_title.strip() else "Ungrouped"
        ws = wb.create_sheet(title=safe_title)
        ws.views.sheetView[0].showGridLines = True

        ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
        ws["A1"] = f"Invoicing Schedule Report ({year} - Months: {month_disp}) - Group: {g} (Generated on: {datetime.now(HKT).strftime('%Y/%m/%d %H:%M')})"
        ws["A1"].font = Font(name="Arial", size=14, bold=True, color="1F497D")
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 30

        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[3].height = 25

        center_cols = ["Client Group", "Incorp Place", "Year", "Anniversary (MM/DD)", "BR No.", "Fee Paid By", "Billing Item", "Fee Deadline", "AR/AFR Deadline", "ES Deadline"]

        group_df = df[df['Client Group'] == g] if 'Client Group' in df.columns else df

        current_row = 4
        for _, item in group_df.iterrows():
            row_dict = {
                "Client Group": item.get("Client Group", ""),
                "Company Name EN": item.get("Company Name EN", ""),
                "Company Name CH": item.get("Company Name CH", ""),
                "Business Name": item.get("Business Name", ""),
                "Incorp Place": item.get("Incorp Place", ""),
                "Year": str(item.get("Year", "")),
                "Anniversary (MM/DD)": item.get("Anniversary (MM/DD)", ""),
                "BR No.": item.get("BR No.", ""),
                "Fee Paid By": item.get("Fee Paid By", ""),
                "Billing Item": item.get("Billing Item", ""),
                "Fee Deadline": item.get("Fee Deadline", ""),
                "AR/AFR Deadline": item.get("AR/AFR Deadline", ""),
                "ES Deadline": item.get("ES Deadline", ""),
                "Remark": item.get("Remark", "")
            }
            
            for col_idx, h in enumerate(headers, 1):
                val = str(row_dict[h])
                cell = ws.cell(row=current_row, column=col_idx, value=clean_status(val) if h == "Billing Item" else val)
                cell.font = font_data
                cell.border = thin_border
                if current_row % 2 == 0: cell.fill = fill_zebra
                
                if h in center_cols: cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else: cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                if h == "Billing Item":
                    cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                    if "Incorp" in val:
                        cell.fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
                    elif "AR + BR" in val or "Package" in val:
                        cell.fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
                    elif "AR Fee Only" in val or "AFR" in val or "ES" in val or "Itemized" in val:
                        cell.fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
                    elif "Branch" in val:
                        cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
                    elif "Cessed" in val:
                        cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                    else:
                        cell.font = font_data
                        if current_row % 2 == 0: cell.fill = fill_zebra
                        else: cell.fill = PatternFill(fill_type=None)
                        
            ws.row_dimensions[current_row].height = 20
            current_row += 1

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row == 1: continue 
                if cell.value: max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(min(max_len + 4, 50), 15)

    if len(wb.sheetnames) == 0:
        wb.create_sheet("No Data")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

@st.cache_data(show_spinner=False)
def generate_inv_pdf(df, year, month_disp, hide_client_group=False):
    now_str = datetime.now(HKT).strftime("%Y/%m/%d %H:%M")
    colspan = "10" if hide_client_group else "11"
    cg_th = "" if hide_client_group else '<th style="width:8%">Client Group</th>'
    
    html = f"""
    <html>
    <head>
        <meta charset="UTF-8">
        <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+TC:wght@400;700&display=swap" rel="stylesheet">
        <style>
            @page {{ 
                size: A4 landscape; margin: 15mm; background-color: #ffffff; 
                @bottom-left {{ content: "Invoicing Schedule Report | Generated on: {now_str}"; font-size: 8pt; color: #7f8c8d; font-family: 'Noto Sans TC', sans-serif; }}
                @bottom-right {{ content: counter(page) " of " counter(pages) " Page(s)"; font-size: 8pt; color: #7f8c8d; font-family: 'Noto Sans TC', sans-serif; }}
            }}
            body {{ font-family: 'Noto Sans TC', sans-serif; font-size: 8pt; color: #2c3e50; margin: 0; padding: 0; }}
            table {{ width: 100%; border-collapse: collapse; }}
            thead {{ display: table-header-group; }}
            tr {{ page-break-inside: avoid; }}
            th {{ background-color: #1f497d; color: white; padding: 5px 3px; border: 1px solid #d9d9d9; font-size: 7.5pt; text-align: center; vertical-align: middle; word-wrap: break-word; }}
            td {{ padding: 5px 3px; border: 1px solid #d9d9d9; text-align: center; vertical-align: middle; word-wrap: break-word; }}
            tr:nth-child(even) td {{ background-color: #f8f9fa; }}
            .text-left {{ text-align: left; font-weight: bold; color: #2980b9; }}
            .report-header-cell {{ border: none !important; background-color: white !important; text-align: left !important; padding: 0 0 10px 0 !important; }}
        </style>
    </head>
    <body>
        <table>
            <thead>
                <tr>
                    <td colspan="{colspan}" class="report-header-cell">
                        <h2 style="color: #1f497d; margin: 0 0 5px 0;">Invoicing Schedule Report ({year} - Months: {month_disp})</h2>
                    </td>
                </tr>
                <tr>
                    {cg_th}
                    <th style="width:14%">Company Name EN</th>
                    <th style="width:8%">Company Name CH</th>
                    <th style="width:5%">Place</th>
                    <th style="width:4%">Year</th>
                    <th style="width:8%">Anniversary<br>(MM/DD)</th>
                    <th style="width:8%">BR No.</th>
                    <th style="width:6%">Fee By</th>
                    <th style="width:12%">Billing Item</th>
                    <th style="width:8%">Fee Deadline</th>
                    <th style="width:8%">AFR/ES Deadline</th>
                    <th>Remark</th>
                </tr>
            </thead>
            <tbody>
    """
    for _, r in df.iterrows():
        cg_td = "" if hide_client_group else f"<td>{r.get('Client Group', '')}</td>"
        
        bill_raw = str(r.get('Billing Item', ''))
        bill_clean = clean_status(bill_raw)
        
        if "Incorp" in bill_raw:
            bill_html = f'<span style="background-color: #9b59b6; color: white; padding: 3px 6px; border-radius: 4px;">{bill_clean}</span>'
        elif "AR + BR" in bill_raw or "Package" in bill_raw:
            bill_html = f'<span style="background-color: #e67e22; color: white; padding: 3px 6px; border-radius: 4px;">{bill_clean}</span>'
        elif "AR Fee Only" in bill_raw or "AFR" in bill_raw or "ES" in bill_raw or "Itemized" in bill_raw:
            bill_html = f'<span style="background-color: #2980b9; color: white; padding: 3px 6px; border-radius: 4px;">{bill_clean}</span>'
        elif "Branch" in bill_raw:
            bill_html = f'<span style="background-color: #27ae60; color: white; padding: 3px 6px; border-radius: 4px;">{bill_clean}</span>'
        elif "Cessed" in bill_raw:
            bill_html = f'<span style="background-color: #e74c3c; color: white; padding: 3px 6px; border-radius: 4px;">{bill_clean}</span>'
        else:
            bill_html = bill_clean
            
        name_en = r.get('Company Name EN', '')
        biz = str(r.get('Business Name', '')).strip()
        if biz and biz not in ['None', 'nan']:
            name_en += f"<br><span style='font-size: 6.5pt; color: #7f8c8d; font-weight: normal;'>Business Name: {biz}</span>"
            
        ar_dl = str(r.get('AR/AFR Deadline', '')).strip()
        es_dl = str(r.get('ES Deadline', '')).strip()
        combo_dl = []
        if ar_dl and ar_dl not in ['N/A', '']: combo_dl.append(f"AFR: {ar_dl}")
        if es_dl and es_dl not in ['N/A', '']: combo_dl.append(f"ES: {es_dl}")
        combo_dl_str = "<br>".join(combo_dl) if combo_dl else "N/A"
        
        html += f"""
        <tr>
            {cg_td}
            <td class="text-left">{name_en}</td>
            <td class="text-left">{r.get('Company Name CH', '')}</td>
            <td>{r.get('Incorp Place', '')}</td>
            <td style="font-weight: bold; color: #1f497d;">{r.get('Year', '')}</td>
            <td style="font-weight: bold;">{r.get('Anniversary (MM/DD)', '')}</td>
            <td style="font-weight: bold;">{r.get('BR No.', '')}</td>
            <td>{r.get('Fee Paid By', '')}</td>
            <td style="font-weight: bold;">{bill_html}</td>
            <td>{r.get('Fee Deadline', '')}</td>
            <td>{combo_dl_str}</td>
            <td style="text-align: left; font-size: 7.5pt; color: #7f8c8d;">{r.get('Remark', '')}</td>
        </tr>
        """
    html += "</tbody></table></body></html>"
    return HTML(string=html).write_pdf()

# --- ZIP Export Logic ---
@st.cache_data(show_spinner=False)
def create_zip_pdfs(df, report_type="All", year=None, month_disp=None):
    now_d = datetime.now(HKT).strftime('%Y%m%d')
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as zip_file:
        col_name = 'Client Group' if 'Client Group' in df.columns else 'client_group'
        groups = df[col_name].unique()
        for g in groups:
            group_df = df[df[col_name] == g]
            safe_g = str(g).replace("/", "_").replace("\\", "_")
            if not safe_g.strip(): safe_g = "Ungrouped"
            
            if report_type == "All":
                pdf_bytes = generate_custom_pdf(group_df, hide_client_group=True)
                filename = f"{safe_g}_Company_Report_{now_d}.pdf"
            elif report_type == "Outstanding":
                pdf_bytes = generate_outstanding_pdf(group_df, hide_client_group=True)
                filename = f"{safe_g}_Outstanding_Report_{now_d}.pdf"
            elif report_type == "Invoicing":
                pdf_bytes = generate_inv_pdf(group_df, year, month_disp, hide_client_group=True)
                filename = f"{safe_g}_Invoicing_Schedule_{year}_{month_disp}_{now_d}.pdf"
            
            zip_file.writestr(filename, pdf_bytes)
    return zip_buffer.getvalue()

@st.cache_data(show_spinner=False)
def create_zip_excels(df, report_type="All", year=None, month_disp=None):
    now_d = datetime.now(HKT).strftime('%Y%m%d')
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as zip_file:
        col_name = 'Client Group' if 'Client Group' in df.columns else 'client_group'
        groups = df[col_name].unique()
        for g in groups:
            group_df = df[df[col_name] == g]
            safe_g = str(g).replace("/", "_").replace("\\", "_")
            if not safe_g.strip(): safe_g = "Ungrouped"
            
            if report_type == "All":
                excel_bytes = generate_general_excel(group_df, hide_client_group=True)
                filename = f"{safe_g}_Company_Report_{now_d}.xlsx"
            elif report_type == "Outstanding":
                excel_bytes = generate_beautiful_excel(group_df, hide_client_group=True)
                filename = f"{safe_g}_Outstanding_Report_{now_d}.xlsx"
            elif report_type == "Invoicing":
                excel_bytes = generate_inv_excel(group_df, year, month_disp, hide_client_group=True)
                filename = f"{safe_g}_Invoicing_Schedule_{year}_{month_disp}_{now_d}.xlsx"
            
            zip_file.writestr(filename, excel_bytes)
    return zip_buffer.getvalue()

# --- 5. Dashboard ---
if choice == "📊 Dashboard":
    st.header("📊 Compliance Overview")
    df_raw = pd.read_sql("SELECT * FROM companies", engine)
    
    df_raw['branch_code'] = df_raw['branch_code'].fillna('000').astype(str).replace(['', 'None', 'nan', '<NA>'], '000')
    df_raw['biz_name'] = df_raw['biz_name'].fillna('').astype(str).replace(['None', 'nan', '<NA>'], '')
    
    groups = pd.read_sql("SELECT group_name FROM client_groups", engine)['group_name'].tolist()
    sorted_groups = sorted([g for g in groups if isinstance(g, str)])
    
    if not df_raw.empty:
        for col in ["incorp_date", "hk_incorp_date", "nd2a_eff_date", "nd4_eff_date", "nd2a_file_date", "nd4_file_date", "nn6_eff_date", "nn6_file_date", "dissolution_date", "ar_ref_date", "br_ref_date", "cessation_date"]:
            if col in df_raw.columns: df_raw[col] = pd.to_datetime(df_raw[col], errors='coerce').dt.date
            
        today = datetime.now(HKT).date()
        outstanding_records = []
        updated_records = []
        
        raw_dict_list = df_raw.to_dict('records')
        companies_with_branches = set(df_raw[df_raw['branch_code'] != '000']['name_en'])
        
        for row in raw_dict_list:
            comp_rec_str = row.get('compliance_records')
            try: comp_rec = json.loads(comp_rec_str) if isinstance(comp_rec_str, str) else {}
            except: comp_rec = {}
            if not isinstance(comp_rec, dict): comp_rec = {}
            
            place = str(row.get('incorp_place', ''))
            is_hk_reg = str(row.get('is_hk_registered', 'False')).strip().lower() in ['true', 'yes', 'y', '1']
            is_bvi = place not in ['HK', ''] and not is_hk_reg
            
            base_date = get_base_date(row)
            if not base_date and is_bvi: base_date = to_date(row.get('incorp_date'))
            incorp_year = base_date.year if base_date else None
            
            cess_date = to_date(row.get('cessation_date'))
            branch_code = str(row.get('branch_code', '000')).strip()
            is_branch = branch_code != '000'
            has_branch = row.get('name_en') in companies_with_branches
            
            prev_br_by = 'Firm'
            prev_afr_fee_by = 'Firm'
            prev_es_fee_by = 'Firm'
            
            for y in active_years:
                y_str = str(y)
                y_data = comp_rec.get(y_str, {})
                
                if incorp_year and y < incorp_year:
                    br_by = 'N/A'
                    afr_fee_by = 'N/A'
                    es_fee_by = 'N/A'
                else:
                    raw_br_by = str(y_data.get('fee_by', y_data.get('br_paid_by', ''))).strip()
                    if raw_br_by: br_by = raw_br_by
                    else: br_by = prev_br_by if prev_br_by != 'N/A' else 'Firm'
                    
                    raw_afr_fee_by = str(y_data.get('afr_fee_by', '')).strip()
                    if raw_afr_fee_by: afr_fee_by = raw_afr_fee_by
                    else: afr_fee_by = prev_afr_fee_by if prev_afr_fee_by != 'N/A' else 'Firm'
                    
                    raw_es_fee_by = str(y_data.get('es_fee_by', '')).strip()
                    if raw_es_fee_by: es_fee_by = raw_es_fee_by
                    else: es_fee_by = prev_es_fee_by if prev_es_fee_by != 'N/A' else 'Firm'
                
                if is_branch and cess_date and y >= cess_date.year:
                    br_by = "N/A"
                    
                prev_br_by = br_by
                prev_afr_fee_by = afr_fee_by
                prev_es_fee_by = es_fee_by
                
                row[f'{y}_br_paid_by'] = br_by
                row[f'{y}_afr_fee_by'] = afr_fee_by
                row[f'{y}_es_fee_by'] = es_fee_by
                
                afr_fee_dt = to_date(y_data.get('afr_fee_date'))
                ar_dt_val = to_date(y_data.get('ar_date'))
                br_dt_val = to_date(y_data.get('fee_date', y_data.get('br_date')))
                es_dt_val = to_date(y_data.get('es_date'))
                es_fee_dt = to_date(y_data.get('es_fee_date'))
                
                if is_branch and cess_date and y >= cess_date.year:
                    br_dt_val = None
                    
                row[f'{y}_br_date'] = br_dt_val
                row[f'{y}_ar_date'] = ar_dt_val
                row[f'{y}_es_date'] = es_dt_val
                row[f'{y}_afr_fee_date'] = afr_fee_dt
                row[f'{y}_es_fee_date'] = es_fee_dt
                
                cr_stat = y_data.get('ar_cr_status', '')
                if not cr_stat: cr_stat = 'Completed' if ar_dt_val else 'Pending'
                row[f'{y}_ar_cr_status'] = cr_stat
                
                es_stat = y_data.get('es_status', '')
                if not es_stat: es_stat = 'Completed' if es_dt_val else 'Pending'
                row[f'{y}_es_status'] = es_stat

            name = str(row.get('name_en', 'Unknown')).strip()
            name_ch = str(row.get('name_ch', ''))
            group = row.get('client_group', '')
            remark_val = str(row.get('remark', ''))
            if remark_val == 'None': remark_val = ""
            
            biz_name = str(row.get('biz_name', '')).strip()
            ar_ref_raw = to_date(row.get('ar_ref_date'))
            br_ref_raw = to_date(row.get('br_ref_date'))
            
            br_no_raw = str(row.get('br_no', '')).strip()
            if has_branch and br_no_raw: disp_br = f"{br_no_raw}-{branch_code}"
            else: disp_br = br_no_raw
            row['disp_br_no'] = disp_br
            
            updated_records.append(row)
            
            if not base_date: continue
            
            for y in report_years:
                y_str = str(y)
                if y < incorp_year: continue 
                
                br_by = str(row.get(f'{y}_br_paid_by', 'Firm'))
                last_br = to_date(row.get(f'{y}_br_date'))
                
                afr_fee_by = str(row.get(f'{y}_afr_fee_by', 'N/A'))
                afr_fee_dt = to_date(row.get(f'{y}_afr_fee_date'))
                
                last_ar = to_date(row.get(f'{y}_ar_date'))
                ar_cr_status = str(row.get(f'{y}_ar_cr_status', 'Pending'))
                
                es_fee_by = str(row.get(f'{y}_es_fee_by', 'N/A'))
                es_fee_dt = to_date(row.get(f'{y}_es_fee_date'))
                last_es = to_date(row.get(f'{y}_es_date'))
                es_status_val = str(row.get(f'{y}_es_status', 'Pending'))
                
                if is_bvi:
                    br_dl = calc_bvi_fee_deadline(base_date, y)
                    ar_dl = calc_afr_deadline(row.get('year_end'), y)
                    es_dl = calc_es_deadline(base_date, y)
                else:
                    if br_ref_raw: br_dl = get_anniv(y, br_ref_raw.month, br_ref_raw.day)
                    else: br_dl = get_anniv(y, base_date.month, base_date.day)
                    ar_dl = calc_ar_deadline(base_date, ar_ref_raw, y)
                    es_dl = None
                
                br_status, ar_status, es_status_disp = "🟢 Normal", "🟢 Normal", "🟢 Normal"
                
                br_dl_str = br_dl.strftime('%Y/%m/%d') if br_dl else "N/A"
                ar_dl_str = ar_dl.strftime('%Y/%m/%d') if ar_dl else "N/A"
                es_dl_str = es_dl.strftime('%Y/%m/%d') if es_dl else "N/A"
                
                br_dt_str = last_br.strftime('%Y/%m/%d') if last_br else ""
                afr_fee_dt_str = afr_fee_dt.strftime('%Y/%m/%d') if afr_fee_dt else ""
                ar_dt_str = last_ar.strftime('%Y/%m/%d') if last_ar else ""
                es_fee_dt_str = es_fee_dt.strftime('%Y/%m/%d') if es_fee_dt else ""
                es_dt_str = last_es.strftime('%Y/%m/%d') if last_es else ""
                
                if is_branch and cess_date and y >= cess_date.year:
                    br_status = "✅ N/A (Cessed)"
                    br_dl_str = "N/A"
                elif br_by == 'Client':
                    if last_br: br_status = "✅ Client (Recorded)"
                    else: br_status = "✅ Client (Pending Record)"
                elif br_by == 'N/A':
                    br_status = "✅ N/A"
                else:
                    if last_br: br_status = "✅ Completed"
                    elif y > today.year: br_status = "🔵 Not yet due"
                    elif br_dl:
                        days_diff = (br_dl - today).days
                        if days_diff < 0: br_status = "🔴 Overdue"
                        elif 0 <= days_diff <= (90 if is_bvi else 30): br_status = "🟡 Due Soon"
                
                if ar_cr_status == 'Exempt (Dormant)':
                    ar_status = "✅ Exempt (Dormant)"
                    ar_dl_str = "N/A"
                elif ar_cr_status == 'Included in Agent Fee':
                    ar_status = "✅ Included"
                    ar_dl_str = "N/A"
                elif is_branch:
                    if cess_date and y >= cess_date.year: ar_status = "✅ N/A (Cessed)"
                    else: ar_status = "✅ N/A (Branch)"
                    ar_dl_str = "N/A"
                elif y == incorp_year:
                    ar_status = "✅ Exempt (1st Year)"
                    ar_dl_str = "N/A"
                else:
                    if last_ar or ar_cr_status == 'Completed': ar_status = "✅ Completed"
                    elif ar_cr_status == 'Processing': ar_status = "⏳ Processing (CR)"
                    elif ar_cr_status == 'Returned': ar_status = "⚠️ Returned (CR)"
                    elif y > today.year: ar_status = "🔵 Not yet due"
                    elif ar_dl:
                        ar_days_diff = (ar_dl - today).days
                        if ar_days_diff < 0: ar_status = "🔴 Overdue"
                        elif 0 <= ar_days_diff <= (90 if is_bvi else 72): ar_status = "🟡 Due Soon"
                
                if is_bvi:
                    if y == incorp_year:
                        es_status_disp = "✅ Exempt (1st Year)"
                        es_dl_str = "N/A"
                    elif es_status_val == 'Exempt': 
                        es_status_disp = "✅ Exempt"
                        es_dl_str = "N/A"
                    elif last_es or es_status_val == 'Completed': es_status_disp = "✅ Completed"
                    elif y > today.year: es_status_disp = "🔵 Not yet due"
                    elif es_dl:
                        es_days_diff = (es_dl - today).days
                        if es_days_diff < 0: es_status_disp = "🔴 Overdue"
                        elif 0 <= es_days_diff <= 90: es_status_disp = "🟡 Due Soon"
                else:
                    es_status_disp = "N/A"
                    es_dl_str = "N/A"
                
                if br_ref_raw and ar_ref_raw and (br_ref_raw.month != ar_ref_raw.month or br_ref_raw.day != ar_ref_raw.day):
                    anniv_disp = f"BR: {br_ref_raw.strftime('%m/%d')} | AR: {ar_ref_raw.strftime('%m/%d')}"
                elif br_ref_raw and not ar_ref_raw and (br_ref_raw.month != base_date.month or br_ref_raw.day != base_date.day):
                    anniv_disp = f"BR: {br_ref_raw.strftime('%m/%d')} | AR: {base_date.strftime('%m/%d')}"
                elif not br_ref_raw and ar_ref_raw and (base_date.month != ar_ref_raw.month or base_date.day != ar_ref_raw.day):
                    anniv_disp = f"BR: {base_date.strftime('%m/%d')} | AR: {ar_ref_raw.strftime('%m/%d')}"
                else:
                    anniv_disp = base_date.strftime('%m/%d')
                        
                # V213: Determine alert if any deadline is approaching or passed, AND date is empty
                is_alert = False
                if not last_br and br_status in ["🔴 Overdue", "🟡 Due Soon"]: is_alert = True
                if not last_ar and ar_status in ["🔴 Overdue", "🟡 Due Soon", "⏳ Processing (CR)", "⚠️ Returned (CR)"]: is_alert = True
                if not last_es and es_status_disp in ["🔴 Overdue", "🟡 Due Soon"]: is_alert = True
                        
                if is_alert:
                    disp_name = f"{name} (-{branch_code})" if has_branch and is_branch else name
                    outstanding_records.append({
                        "Company Name EN": disp_name,
                        "Company Name CH": name_ch,
                        "Business Name": biz_name,
                        "Client Group": group,
                        "Incorp Place": place,
                        "Year": y_str,
                        "Anniversary (MM/DD)": anniv_disp,
                        "BR No.": disp_br,
                        "Fee Paid By": br_by,
                        "Fee Paid Date": br_dt_str,
                        "Fee Deadline": br_dl_str,
                        "Fee Status": br_status,
                        "AR/AFR Fee By": afr_fee_by if is_bvi else "N/A",
                        "AR/AFR Fee Date": afr_fee_dt_str if is_bvi else "N/A",
                        "AR/AFR Filed Date": ar_dt_str,
                        "AR/AFR Deadline": ar_dl_str,
                        "AR/AFR Status": ar_status,
                        "ES Fee By": es_fee_by if is_bvi else "N/A",
                        "ES Fee Date": es_fee_dt_str if is_bvi else "N/A",
                        "ES Filed Date": es_dt_str if is_bvi else "N/A",
                        "ES Deadline": es_dl_str,
                        "ES Status": es_status_disp,
                        "Remark": remark_val,
                        "branch_code_raw": branch_code
                    })

        df_raw = pd.DataFrame(updated_records)

        tab1, tab2, tab3 = st.tabs(["📊 All Companies", "🚨 Outstanding List", "🧾 Invoicing Schedule"])
        
        with tab1:
            sort_cols = [c for c in ['client_group', 'name_en', 'branch_code', 'incorp_place'] if c in df_raw.columns]
            df_raw = df_raw.sort_values(by=sort_cols, na_position='last')
            
            # V213: Year Filter Integration (±1 year logic)
            t1, t2, t3, t4, t5 = st.columns([2, 2, 2, 2, 4])
            filter_g = t1.selectbox("🔍 Filter Group", ["All Groups"] + sorted_groups)
            target_year_disp = t2.selectbox("📅 Display Target Year", active_years, index=active_years.index(current_system_year))
            
            if t3.button("🔄 Refresh"): st.rerun()
            if 'sel_v213' not in st.session_state: st.session_state.sel_v213 = False
            if t4.button("✅ Select All"): st.session_state.sel_v213 = True; st.rerun()
            if t5.button("🧹 Clear All"): st.session_state.sel_v213 = False; st.rerun()
            
            disp_years = [y for y in active_years if target_year_disp - 1 <= y <= target_year_disp + 1]
            
            df_filtered = df_raw if filter_g == "All Groups" else df_raw[df_raw['client_group'] == filter_g]
            
            display_cols_ordered = [
                "name_ch", "biz_name", "client_group", "incorp_place", "incorp_place_others", 
                "incorp_date", "ci_no", "is_hk_registered", "hk_incorp_date", "hk_ci_no", 
                "disp_br_no", "co_type", "agent", "year_end", "billing_mode", "reg_addr", "corres_addr", "round_loc", "sign_loc", "seal_loc"
            ]
            
            base_cols = [c for c in TEMPLATE_COLS if c in df_filtered.columns]
            for c in base_cols:
                if c not in display_cols_ordered and c not in ["remark", "name_en", "branch_code", "ar_ref_date", "br_ref_date", "cessation_date", "br_no"]: 
                    display_cols_ordered.append(c)
            
            dyn_cols = []
            for y in disp_years:
                dyn_cols.extend([f"{y}_br_paid_by", f"{y}_br_date", f"{y}_afr_fee_by", f"{y}_afr_fee_date", f"{y}_ar_date", f"{y}_ar_cr_status", f"{y}_es_fee_by", f"{y}_es_fee_date", f"{y}_es_date", f"{y}_es_status"])
            
            display_cols_ordered.extend(dyn_cols)
            display_cols_ordered.append("remark")
            display_cols_ordered.append("branch_code")
            
            df_display = df_filtered[["name_en"] + [c for c in display_cols_ordered if c in df_filtered.columns]].copy()
            
            if not df_display.empty:
                df_display['remark'] = df_display['remark'].fillna('')
                
                def format_name(r):
                    if str(r['branch_code']).strip() != '000' and r['name_en'] in companies_with_branches:
                        return f"{r['name_en']} (-{str(r['branch_code']).strip()})"
                    return r['name_en']
                
                df_display['name_en'] = df_display.apply(format_name, axis=1)
                
                df_display.rename(columns=EXCHANGE_COL_MAPPING, inplace=True, errors='ignore')
                df_display.rename(columns={'disp_br_no': 'BR No.'}, inplace=True, errors='ignore')
                
                dyn_rename_dict = {}
                for y in disp_years:
                    dyn_rename_dict[f"{y}_br_paid_by"] = f"{y} Fee Paid By"
                    dyn_rename_dict[f"{y}_br_date"] = f"{y} Fee Paid Date"
                    dyn_rename_dict[f"{y}_afr_fee_by"] = f"{y} AR/AFR Fee Paid By"
                    dyn_rename_dict[f"{y}_afr_fee_date"] = f"{y} AR/AFR Fee Paid Date"
                    dyn_rename_dict[f"{y}_ar_date"] = f"{y} AR/AFR Filed Date"
                    dyn_rename_dict[f"{y}_ar_cr_status"] = f"{y} AR/AFR Status"
                    dyn_rename_dict[f"{y}_es_fee_by"] = f"{y} ES Fee Paid By"
                    dyn_rename_dict[f"{y}_es_fee_date"] = f"{y} ES Fee Paid Date"
                    dyn_rename_dict[f"{y}_es_date"] = f"{y} ES Filed Date"
                    dyn_rename_dict[f"{y}_es_status"] = f"{y} ES Status"
                df_display.rename(columns=dyn_rename_dict, inplace=True)
                
                df_display.insert(0, "Select", st.session_state.sel_v213)
                
                s = df_display["Company Name EN"].astype(str)
                df_display.index = s + s.groupby(s).cumcount().map(lambda x: '\u200B' * x)
                df_display.index.name = "Company Name EN"
                df_display.drop(columns=["Company Name EN"], inplace=True)
                
                st.markdown(f"📈 Total: **{len(df_filtered)}** records. Only showing columns for years: **{', '.join(map(str, disp_years))}**.")
                
                col_cfg = {
                    "Select": st.column_config.CheckboxColumn("Select", default=False),
                    "Branch Code": None, 
                    "Remark": st.column_config.TextColumn("Remark")
                }
                cr_opts = ["Pending", "Processing", "Returned", "Completed", "Exempt (Dormant)", "Included in Agent Fee"]
                pay_opts = ["Firm", "Client", "N/A"]
                es_opts = ["Pending", "Completed", "Exempt"]
                for y in disp_years:
                    col_cfg[f"{y} Fee Paid By"] = st.column_config.SelectboxColumn(f"✏️ {y} Fee Paid By", options=pay_opts, required=True)
                    col_cfg[f"{y} Fee Paid Date"] = st.column_config.DateColumn(f"✏️ {y} Fee Paid Date", format="YYYY/MM/DD")
                    col_cfg[f"{y} AR/AFR Fee Paid By"] = st.column_config.SelectboxColumn(f"✏️ {y} AR/AFR Fee Paid By", options=pay_opts, required=True)
                    col_cfg[f"{y} AR/AFR Fee Paid Date"] = st.column_config.DateColumn(f"✏️ {y} AR/AFR Fee Paid Date", format="YYYY/MM/DD")
                    col_cfg[f"{y} AR/AFR Filed Date"] = st.column_config.DateColumn(f"✏️ {y} AR/AFR Filed Date", format="YYYY/MM/DD")
                    col_cfg[f"{y} AR/AFR Status"] = st.column_config.SelectboxColumn(f"✏️ {y} AR/AFR Status", options=cr_opts, required=True)
                    col_cfg[f"{y} ES Fee Paid By"] = st.column_config.SelectboxColumn(f"✏️ {y} ES Fee Paid By", options=pay_opts, required=True)
                    col_cfg[f"{y} ES Fee Paid Date"] = st.column_config.DateColumn(f"✏️ {y} ES Fee Paid Date", format="YYYY/MM/DD")
                    col_cfg[f"{y} ES Filed Date"] = st.column_config.DateColumn(f"✏️ {y} ES Filed Date", format="YYYY/MM/DD")
                    col_cfg[f"{y} ES Status"] = st.column_config.SelectboxColumn(f"✏️ {y} ES Status", options=es_opts, required=True)
                
                disabled_cols = [c for c in df_display.columns if c not in ["Select", "Remark"] and not any(c.endswith(suffix) for suffix in ["Fee Paid By", "Fee Paid Date", "AR/AFR Fee Paid By", "AR/AFR Fee Paid Date", "AR/AFR Filed Date", "AR/AFR Status", "ES Fee Paid By", "ES Fee Paid Date", "ES Filed Date", "ES Status"])]
                
                edit_df = st.data_editor(
                    df_display, 
                    column_config=col_cfg,
                    disabled=disabled_cols,
                    use_container_width=True,
                    key="dash_v213"
                )
                
                if st.button("💾 Save Batch Edits", key="btn_save_grid_v213"):
                    try:
                        with engine.begin() as conn:
                            for c_name_idx, r in edit_df.iterrows():
                                c_n = str(c_name_idx).replace('\u200B', '')
                                b_code = str(r['Branch Code'])
                                suffix = f" (-{b_code})"
                                if c_n.endswith(suffix):
                                    c_n = c_n[:-len(suffix)]
                                
                                row_info = df_raw[(df_raw['name_en'] == c_n) & (df_raw['branch_code'] == b_code)].iloc[0]
                                
                                comp_name_safe = str(row_info['name_en']).replace("'", "''")
                                branch_safe = str(row_info['branch_code']).replace("'", "''")
                                
                                base_dt = get_base_date(row_info)
                                inc_yr = base_dt.year if base_dt else None
                                is_bvi = str(row_info['incorp_place']) not in ['HK', ''] and str(row_info['is_hk_registered']).lower() not in ['true', 'yes', '1']
                                b_mode = str(row_info.get('billing_mode', 'Itemized'))
                                
                                try: comp_dict_existing = json.loads(row_info.get('compliance_records', '{}'))
                                except: comp_dict_existing = {}
                                if not isinstance(comp_dict_existing, dict): comp_dict_existing = {}
                                
                                comp_dict = {}
                                prev_br_by = 'Firm'
                                prev_afr_fee_by = 'Firm'
                                prev_es_fee_by = 'Firm'
                                
                                for y in active_years:
                                    y_str = str(y)
                                    
                                    if y in disp_years:
                                        raw_br_by = str(r.get(f'{y} Fee Paid By', '')).strip()
                                        raw_afr = str(r.get(f'{y} AR/AFR Fee Paid By', '')).strip()
                                        raw_es = str(r.get(f'{y} ES Fee Paid By', '')).strip()
                                        
                                        in_ar_cr = str(r.get(f'{y} AR/AFR Status', '')).strip()
                                        in_es_st = str(r.get(f'{y} ES Status', '')).strip()
                                        
                                        br_date_val = r.get(f'{y} Fee Paid Date')
                                        afr_fee_date_val = r.get(f'{y} AR/AFR Fee Paid Date')
                                        ar_date_val = r.get(f'{y} AR/AFR Filed Date')
                                        es_fee_date_val = r.get(f'{y} ES Fee Paid Date')
                                        es_date_val = r.get(f'{y} ES Filed Date')
                                    else:
                                        db_y_data = comp_dict_existing.get(y_str, {})
                                        raw_br_by = str(db_y_data.get('fee_by', db_y_data.get('br_paid_by', ''))).strip()
                                        raw_afr = str(db_y_data.get('afr_fee_by', '')).strip()
                                        raw_es = str(db_y_data.get('es_fee_by', '')).strip()
                                        
                                        in_ar_cr = str(db_y_data.get('ar_cr_status', 'Pending')).strip()
                                        in_es_st = str(db_y_data.get('es_status', 'Pending')).strip()
                                        
                                        br_date_val = db_y_data.get('fee_date', db_y_data.get('br_date'))
                                        afr_fee_date_val = db_y_data.get('afr_fee_date')
                                        ar_date_val = db_y_data.get('ar_date')
                                        es_fee_date_val = db_y_data.get('es_fee_date')
                                        es_date_val = db_y_data.get('es_date')
                                    
                                    br_by = raw_br_by if raw_br_by else (prev_br_by if prev_br_by != 'N/A' else 'Firm')
                                    
                                    new_ar_cr = in_ar_cr if in_ar_cr in ["Pending", "Processing", "Returned", "Completed", "Exempt (Dormant)", "Included in Agent Fee"] else "Pending"
                                    new_es_st = in_es_st if in_es_st in ["Pending", "Completed", "Exempt"] else "Pending"
                                    
                                    if is_bvi and 'All-in' in b_mode:
                                        afr_fee_by = br_by
                                        raw_afr_fee = to_date(br_date_val)
                                    else:
                                        afr_fee_by = raw_afr if raw_afr else (prev_afr_fee_by if prev_afr_fee_by != 'N/A' else 'Firm')
                                        raw_afr_fee = to_date(afr_fee_date_val)
                                        
                                    es_fee_by = raw_es if raw_es else (prev_es_fee_by if prev_es_fee_by != 'N/A' else 'Firm')
                                    
                                    prev_br_by = br_by
                                    prev_afr_fee_by = afr_fee_by
                                    prev_es_fee_by = es_fee_by
                                    
                                    raw_br = to_date(br_date_val)
                                    raw_ar = to_date(ar_date_val)
                                    raw_es_fee = to_date(es_fee_date_val)
                                    raw_es = to_date(es_date_val)
                                    
                                    if inc_yr and y < inc_yr:
                                        br_by, afr_fee_by, es_fee_by = 'N/A', 'N/A', 'N/A'
                                        raw_br, raw_afr_fee, raw_ar, raw_es_fee, raw_es = None, None, None, None, None
                                    elif inc_yr and y == inc_yr:
                                        raw_ar, raw_es = None, None
                                        
                                    if br_by == 'N/A': raw_br = None
                                    if afr_fee_by == 'N/A': raw_afr_fee = None
                                    if es_fee_by == 'N/A': raw_es_fee = None
                                    if new_ar_cr in ['Exempt (Dormant)', 'Included in Agent Fee']: raw_ar = None
                                    if new_es_st == 'Exempt': raw_es = None
                                        
                                    comp_dict[y_str] = {
                                        "br_paid_by": br_by,
                                        "br_date": raw_br.strftime('%Y-%m-%d') if raw_br else None,
                                        "afr_fee_by": afr_fee_by,
                                        "afr_fee_date": raw_afr_fee.strftime('%Y-%m-%d') if raw_afr_fee else None,
                                        "ar_date": raw_ar.strftime('%Y-%m-%d') if raw_ar else None,
                                        "ar_cr_status": new_ar_cr,
                                        "es_fee_by": es_fee_by,
                                        "es_fee_date": raw_es_fee.strftime('%Y-%m-%d') if raw_es_fee else None,
                                        "es_date": raw_es.strftime('%Y-%m-%d') if raw_es else None,
                                        "es_status": new_es_st
                                    }
                                json_str = json.dumps(comp_dict).replace("'", "''")
                                rem_str = str(r.get('Remark', '')).replace("'", "''")
                                if rem_str == 'None': rem_str = ""
                                
                                sql = f"UPDATE companies SET compliance_records = '{json_str}', remark = '{rem_str}' WHERE name_en = '{comp_name_safe}' AND branch_code = '{branch_safe}'"
                                conn.execute(text(sql))
                        st.success("✅ Changes saved successfully!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Batch save failed: {e}")
                
                selected = edit_df[edit_df["Select"] == True]
                if len(selected) > 0:
                    st.info(f"✅ **{len(selected)}** records selected for action.")
                    
                    selected_tuples = []
                    for c_n_idx, r in selected.iterrows():
                        c_n = str(c_n_idx).replace('\u200B', '')
                        b_code = str(r['Branch Code'])
                        suffix = f" (-{b_code})"
                        if c_n.endswith(suffix):
                            c_n = c_n[:-len(suffix)]
                        selected_tuples.append((c_n, b_code))
                    
                    def match_selected(r):
                        name = str(r['name_en']).strip()
                        bcode = str(r['branch_code']).strip()
                        return (name, bcode) in selected_tuples
                        
                    mask = df_raw.apply(match_selected, axis=1)
                    final_data = df_raw[mask]
                    
                    c_act1, c_act2, c_act3 = st.columns([4, 2, 2])
                    
                    exp_opt1 = c_act1.selectbox(
                        "📤 Select Export Format",
                        [
                            "🏢 Internal Export - PDF (Combined)",
                            "🏢 Internal Export - Excel (Combined)",
                            "🤝 Client Export - ZIP (PDFs by Group)",
                            "🤝 Client Export - ZIP (Excels by Group)",
                            "📄 External Export - PDF (No Group)",
                            "📄 External Export - Excel (No Group)"
                        ],
                        label_visibility="collapsed",
                        key="sel_exp1"
                    )
                    
                    if c_act2.button("🚀 Generate Report", use_container_width=True, key="btn_gen1"):
                        with st.spinner("Generating... Please wait..."):
                            now_d = datetime.now(HKT).strftime('%Y%m%d')
                            if "Internal Export - PDF" in exp_opt1:
                                st.session_state.ex_data1 = generate_custom_pdf(final_data)
                                st.session_state.ex_name1 = f"Company_Report_Internal_{now_d}.pdf"
                                st.session_state.ex_mime1 = "application/pdf"
                            elif "Internal Export - Excel" in exp_opt1:
                                st.session_state.ex_data1 = generate_general_excel(final_data)
                                st.session_state.ex_name1 = f"Company_Report_Internal_{now_d}.xlsx"
                                st.session_state.ex_mime1 = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            elif "Client Export - ZIP (PDFs" in exp_opt1:
                                st.session_state.ex_data1 = create_zip_pdfs(final_data, "All")
                                st.session_state.ex_name1 = f"Company_Reports_PDF_{now_d}.zip"
                                st.session_state.ex_mime1 = "application/zip"
                            elif "Client Export - ZIP (Excels" in exp_opt1:
                                st.session_state.ex_data1 = create_zip_excels(final_data, "All")
                                st.session_state.ex_name1 = f"Company_Reports_Excel_{now_d}.zip"
                                st.session_state.ex_mime1 = "application/zip"
                            elif "External Export - PDF" in exp_opt1:
                                st.session_state.ex_data1 = generate_custom_pdf(final_data, hide_client_group=True)
                                st.session_state.ex_name1 = f"Company_Report_External_{now_d}.pdf"
                                st.session_state.ex_mime1 = "application/pdf"
                            elif "External Export - Excel" in exp_opt1:
                                st.session_state.ex_data1 = generate_general_excel(final_data, hide_client_group=True)
                                st.session_state.ex_name1 = f"Company_Report_External_{now_d}.xlsx"
                                st.session_state.ex_mime1 = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

                    if 'ex_data1' in st.session_state:
                        c_act3.download_button(
                            "📥 Click to Download", 
                            data=st.session_state.ex_data1, 
                            file_name=st.session_state.ex_name1, 
                            mime=st.session_state.ex_mime1,
                            type="primary", 
                            use_container_width=True, 
                            key="btn_dl1"
                        )
                        
                    st.write("---")
                    with st.popover("🧨 BATCH DELETE"):
                        st.error("🛑 DANGER ZONE")
                        conf_b = st.text_input("Type DELETE", key="batch_del_v213")
                        if st.button("Confirm Batch Delete", disabled=(conf_b != "DELETE"), key="btn_batch_del_v213"):
                            df_raw[~mask].to_sql('companies', engine, if_exists='replace', index=False)
                            st.rerun()
            else:
                st.info("No records match the current filter.")

        with tab2:
            df_alerts = pd.DataFrame(outstanding_records)
            if not df_alerts.empty:
                df_alerts = df_alerts.sort_values(by=['Client Group', 'Company Name EN', 'branch_code_raw', 'Incorp Place', 'Year'], na_position='last')
                
                ta1, ta2, ta3, ta4 = st.columns([3, 2, 2, 5])
                filter_alert_g = ta1.selectbox("🔍 Filter Group", ["All Groups"] + sorted_groups, key="filter_alert_g")
                if ta2.button("🔄 Refresh", key="ref_alert"): st.rerun()
                
                df_alerts_filtered = df_alerts if filter_alert_g == "All Groups" else df_alerts[df_alerts['Client Group'] == filter_alert_g]
                
                if 'sel_alert_v213' not in st.session_state: st.session_state.sel_alert_v213 = False
                if ta3.button("✅ Select All", key="sel_all_alert"): st.session_state.sel_alert_v213 = True; st.rerun()
                if ta4.button("🧹 Clear All", key="clr_all_alert"): st.session_state.sel_alert_v213 = False; st.rerun()
                
                # V213: Restored Status columns in Outstanding List
                alert_cols_order = ["Company Name EN", "Company Name CH", "Business Name", "Client Group", "Incorp Place", "Year", "Anniversary (MM/DD)", "BR No.", "Fee Paid By", "Fee Paid Date", "Fee Deadline", "Fee Status", "AR/AFR Fee By", "AR/AFR Fee Date", "AR/AFR Filed Date", "AR/AFR Deadline", "AR/AFR Status", "ES Fee By", "ES Fee Date", "ES Filed Date", "ES Deadline", "ES Status", "Remark", "branch_code_raw"]
                df_alerts_display = df_alerts_filtered[alert_cols_order].copy()
                
                if not df_alerts_display.empty:
                    df_alerts_display.insert(0, "Select", st.session_state.sel_alert_v213)
                    
                    s2 = df_alerts_display["Company Name EN"].astype(str)
                    df_alerts_display.index = s2 + s2.groupby(s2).cumcount().map(lambda x: '\u200B' * x)
                    df_alerts_display.index.name = "Company Name EN"
                    df_alerts_display.drop(columns=["Company Name EN"], inplace=True)
                    
                    st.markdown(f"📈 Total: **{len(df_alerts_display)}** tasks in current view.")
                    
                    alert_edit = st.data_editor(
                        df_alerts_display,
                        column_config={
                            "Select": st.column_config.CheckboxColumn("Select", default=False),
                            "branch_code_raw": None
                        }, 
                        use_container_width=True,
                        disabled=[c for c in df_alerts_display.columns if c != "Select"],
                        key="alert_grid_v213"
                    )
                    
                    selected_alerts = df_alerts_display[alert_edit["Select"] == True]
                    sel_alert_tuples = []
                    for c_n_idx, r in selected_alerts.iterrows():
                        c_n = str(c_n_idx).replace('\u200B', '')
                        sel_alert_tuples.append((c_n, r['branch_code_raw']))
                        
                    def match_alert_selected(r):
                        name = str(r['Company Name EN']).strip()
                        bcode = str(r['branch_code_raw']).strip()
                        suffix = f" (-{bcode})"
                        if name.endswith(suffix):
                            name = name[:-len(suffix)]
                        return (name, bcode) in sel_alert_tuples
                        
                    export_target = df_alerts[df_alerts.apply(match_alert_selected, axis=1)]
                    
                    if len(export_target) > 0:
                        export_target = export_target.drop(columns=['branch_code_raw'])
                    else: export_target = None
                    
                    if export_target is not None:
                        st.info(f"✅ **{len(export_target)}** tasks selected for export.")
                        
                        ca1, ca2, ca3 = st.columns([4, 2, 2])
                        exp_opt2 = ca1.selectbox(
                            "📤 Select Export Format",
                            [
                                "🏢 Internal Export - PDF (Combined)",
                                "🏢 Internal Export - Excel (Combined)",
                                "🤝 Client Export - ZIP (PDFs by Group)",
                                "🤝 Client Export - ZIP (Excels by Group)",
                                "📄 External Export - PDF (No Group)",
                                "📄 External Export - Excel (No Group)"
                            ],
                            label_visibility="collapsed",
                            key="sel_exp2"
                        )
                        
                        if ca2.button("🚀 Generate Report", use_container_width=True, key="btn_gen2"):
                            with st.spinner("Generating... Please wait..."):
                                now_d = datetime.now(HKT).strftime('%Y%m%d')
                                if "Internal Export - PDF" in exp_opt2:
                                    st.session_state.ex_data2 = generate_outstanding_pdf(export_target)
                                    st.session_state.ex_name2 = f"Outstanding_Report_Internal_{now_d}.pdf"
                                    st.session_state.ex_mime2 = "application/pdf"
                                elif "Internal Export - Excel" in exp_opt2:
                                    st.session_state.ex_data2 = generate_beautiful_excel(export_target)
                                    st.session_state.ex_name2 = f"Outstanding_Report_Internal_{now_d}.xlsx"
                                    st.session_state.ex_mime2 = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                elif "Client Export - ZIP (PDFs" in exp_opt2:
                                    st.session_state.ex_data2 = create_zip_pdfs(export_target, "Outstanding")
                                    st.session_state.ex_name2 = f"Outstanding_PDF_{now_d}.zip"
                                    st.session_state.ex_mime2 = "application/zip"
                                elif "Client Export - ZIP (Excels" in exp_opt2:
                                    st.session_state.ex_data2 = create_zip_excels(export_target, "Outstanding")
                                    st.session_state.ex_name2 = f"Outstanding_Excel_{now_d}.zip"
                                    st.session_state.ex_mime2 = "application/zip"
                                elif "External Export - PDF" in exp_opt2:
                                    st.session_state.ex_data2 = generate_outstanding_pdf(export_target, hide_client_group=True)
                                    st.session_state.ex_name2 = f"Outstanding_Report_External_{now_d}.pdf"
                                    st.session_state.ex_mime2 = "application/pdf"
                                elif "External Export - Excel" in exp_opt2:
                                    st.session_state.ex_data2 = generate_beautiful_excel(export_target, hide_client_group=True)
                                    st.session_state.ex_name2 = f"Outstanding_Report_External_{now_d}.xlsx"
                                    st.session_state.ex_mime2 = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

                        if 'ex_data2' in st.session_state:
                            ca3.download_button(
                                "📥 Click to Download", 
                                data=st.session_state.ex_data2, 
                                file_name=st.session_state.ex_name2, 
                                mime=st.session_state.ex_mime2,
                                type="primary", 
                                use_container_width=True, 
                                key="btn_dl2"
                            )
                    else:
                        st.info("💡 Select checkboxes to export specific companies.")
                else:
                    st.info("No outstanding tasks for this Group.")
            else:
                st.success("🎉 No outstanding tasks at the moment!")
                
        with tab3:
            ti1, ti2, ti3, ti4, ti5, ti6, ti7 = st.columns([2, 1.5, 1.5, 1.5, 1, 1, 3])
            filter_inv_g = ti1.selectbox("🔍 Filter Group", ["All Groups"] + sorted_groups, key="filter_inv_g")
            
            target_year = ti2.selectbox("📅 Target Year", active_years, index=active_years.index(current_system_year), key="inv_year")
            
            start_month = ti3.selectbox("📆 From Month", range(1, 13), index=0, key="inv_start_m")
            end_month = ti4.selectbox("📆 To Month", range(1, 13), index=11, key="inv_end_m")
            month_str = f"{start_month}-{end_month}" if start_month != end_month else str(start_month)
            
            if ti5.button("🔄 Refresh", key="ref_inv"): st.rerun()
            
            if 'sel_inv_v213' not in st.session_state: st.session_state.sel_inv_v213 = False
            if ti6.button("✅ Select All", key="sel_all_inv"): st.session_state.sel_inv_v213 = True; st.rerun()
            if ti7.button("🧹 Clear All", key="clr_all_inv"): st.session_state.sel_inv_v213 = False; st.rerun()
            
            inv_records = []
            for row in raw_dict_list:
                place = str(row.get('incorp_place', ''))
                is_hk_reg = str(row.get('is_hk_registered', 'False')).strip().lower() in ['true', 'yes', 'y', '1']
                is_bvi = place not in ['HK', ''] and not is_hk_reg
                
                base_date = get_base_date(row)
                if not base_date and is_bvi: base_date = to_date(row.get('incorp_date'))
                if not base_date: continue
                incorp_year = base_date.year
                
                if target_year < incorp_year: continue
                
                br_ref_raw = to_date(row.get('br_ref_date'))
                ar_ref_raw = to_date(row.get('ar_ref_date'))
                
                if is_bvi:
                    br_dl = calc_bvi_fee_deadline(base_date, target_year)
                    ar_dl = calc_afr_deadline(row.get('year_end'), target_year)
                    es_dl = calc_es_deadline(base_date, target_year)
                    
                    inv_targets = []
                    if br_dl: inv_targets.append(add_months(br_dl, -3).month)
                    if ar_dl: inv_targets.append(add_months(ar_dl, -3).month)
                    if es_dl: inv_targets.append(add_months(es_dl, -3).month)
                    
                    matched = False
                    for m in set(inv_targets):
                        if start_month <= end_month:
                            if start_month <= m <= end_month: matched = True
                        else: 
                            if m >= start_month or m <= end_month: matched = True
                    if not matched: continue
                else:
                    filter_date = br_ref_raw if br_ref_raw else base_date
                    b_month = filter_date.month
                    if start_month <= end_month:
                        if not (start_month <= b_month <= end_month): continue
                    else: 
                        if not (b_month >= start_month or b_month <= end_month): continue
                
                group = row.get('client_group', '')
                if filter_inv_g != "All Groups" and group != filter_inv_g: continue
                
                name = str(row.get('name_en', 'Unknown')).strip()
                name_ch = str(row.get('name_ch', ''))
                remark_val = str(row.get('remark', ''))
                if remark_val == 'None': remark_val = ""
                
                biz_name = str(row.get('biz_name', '')).strip()
                branch_code = str(row.get('branch_code', '000')).strip()
                is_branch = branch_code != '000'
                cess_date = to_date(row.get('cessation_date'))
                
                br_by = str(row.get(f'{target_year}_br_paid_by', 'Firm'))
                if is_branch and cess_date and target_year >= cess_date.year:
                    br_by = "N/A"
                
                billing_mode = str(row.get('billing_mode', 'Itemized'))
                if 'All-in' in billing_mode: billing_mode = 'All-in Package'
                elif 'Itemized' in billing_mode: billing_mode = 'Itemized'
                
                if is_branch:
                    if target_year == incorp_year: billing_item = "🆕 Branch Registration"
                    elif cess_date and target_year >= cess_date.year: billing_item = "⛔ Cessed"
                    elif br_by == "Firm": billing_item = "💰 BR Fee Only (Branch)"
                    else: billing_item = "✅ N/A (Client Paid)"
                elif is_bvi:
                    if target_year == incorp_year: billing_item = "🆕 BVI Incorp. Package"
                    else:
                        if 'All-in' in billing_mode: billing_item = "💰 BVI Annual Package (Fee+AFR)"
                        else: billing_item = "📄 BVI Itemized (Fee / AFR / ES)"
                else:
                    if target_year == incorp_year: billing_item = "🆕 Incorp. Package"
                    elif br_by == "Firm": billing_item = "💰 AR + BR Fee"
                    else: billing_item = "📄 AR Fee Only"
                
                if not is_bvi:
                    if br_ref_raw: br_dl = get_anniv(target_year, br_ref_raw.month, br_ref_raw.day)
                    else: br_dl = get_anniv(target_year, base_date.month, base_date.day)
                    ar_dl = calc_ar_deadline(base_date, ar_ref_raw, target_year)
                    es_dl = None
                
                br_dl_str = br_dl.strftime('%Y/%m/%d') if br_dl else "N/A"
                if is_branch and cess_date and target_year >= cess_date.year:
                    br_dl_str = "N/A"
                
                ar_cr_status_val = str(row.get(f'{target_year}_ar_cr_status', 'Pending'))
                if ar_cr_status_val == 'Exempt (Dormant)': ar_dl_str = "Exempt (Dormant)"
                elif ar_cr_status_val == 'Included in Agent Fee': ar_dl_str = "Included"
                elif is_branch: ar_dl_str = "N/A"
                elif target_year == incorp_year: ar_dl_str = "Exempt (1st Year)"
                else: ar_dl_str = ar_dl.strftime('%Y/%m/%d') if ar_dl else "N/A"
                    
                if not is_bvi: es_dl_str = "N/A"
                elif target_year == incorp_year: es_dl_str = "Exempt"
                else: es_dl_str = es_dl.strftime('%Y/%m/%d') if es_dl else "N/A"
                    
                disp_name = f"{name} (-{branch_code})" if has_branch and is_branch else name
                
                if br_ref_raw and ar_ref_raw and (br_ref_raw.month != ar_ref_raw.month or br_ref_raw.day != ar_ref_raw.day):
                    anniv_disp = f"BR: {br_ref_raw.strftime('%m/%d')} | AR: {ar_ref_raw.strftime('%m/%d')}"
                elif br_ref_raw and not ar_ref_raw and (br_ref_raw.month != base_date.month or br_ref_raw.day != base_date.day):
                    anniv_disp = f"BR: {br_ref_raw.strftime('%m/%d')} | AR: {base_date.strftime('%m/%d')}"
                elif not br_ref_raw and ar_ref_raw and (base_date.month != ar_ref_raw.month or base_date.day != ar_ref_raw.day):
                    anniv_disp = f"BR: {base_date.strftime('%m/%d')} | AR: {ar_ref_raw.strftime('%m/%d')}"
                else:
                    anniv_disp = base_date.strftime('%m/%d')
                    
                inv_records.append({
                    "Company Name EN": disp_name,
                    "Company Name CH": name_ch,
                    "Business Name": biz_name,
                    "Client Group": group,
                    "Incorp Place": place,
                    "Year": str(target_year),
                    "Anniversary (MM/DD)": anniv_disp,
                    "BR No.": row.get('disp_br_no', ''),
                    "Fee Paid By": br_by,
                    "Billing Item": billing_item,
                    "Fee Deadline": br_dl_str,
                    "AR/AFR Deadline": ar_dl_str,
                    "ES Deadline": es_dl_str,
                    "Remark": remark_val,
                    "branch_code_raw": branch_code
                })
                
            df_inv = pd.DataFrame(inv_records)
            if not df_inv.empty:
                df_inv = df_inv.sort_values(by=['Client Group', 'Company Name EN', 'branch_code_raw', 'Incorp Place', 'Year', 'Anniversary (MM/DD)'])
                
                inv_cols_order = ["Company Name EN", "Company Name CH", "Business Name", "Client Group", "Incorp Place", "Year", "Anniversary (MM/DD)", "BR No.", "Fee Paid By", "Billing Item", "Fee Deadline", "AR/AFR Deadline", "ES Deadline", "Remark", "branch_code_raw"]
                df_inv_display = df_inv[inv_cols_order].copy()
                
                if not df_inv_display.empty:
                    df_inv_display.insert(0, "Select", st.session_state.sel_inv_v213)
                    
                    s3 = df_inv_display["Company Name EN"].astype(str)
                    df_inv_display.index = s3 + s3.groupby(s3).cumcount().map(lambda x: '\u200B' * x)
                    df_inv_display.index.name = "Company Name EN"
                    df_inv_display.drop(columns=["Company Name EN"], inplace=True)
                    
                    st.markdown(f"📈 Total: **{len(df_inv_display)}** records for Invoicing in current view.")
                    
                    inv_edit = st.data_editor(
                        df_inv_display,
                        column_config={
                            "Select": st.column_config.CheckboxColumn("Select", default=False),
                            "branch_code_raw": None
                        }, 
                        use_container_width=True,
                        disabled=[c for c in df_inv_display.columns if c != "Select"],
                        key="inv_grid_v213"
                    )
                    
                    selected_inv = df_inv_display[inv_edit["Select"] == True]
                    sel_inv_tuples = []
                    for c_n_idx, r in selected_inv.iterrows():
                        c_n = str(c_n_idx).replace('\u200B', '')
                        sel_inv_tuples.append((c_n, r['branch_code_raw']))
                        
                    def match_inv_selected(r):
                        name = str(r['Company Name EN']).strip()
                        bcode = str(r['branch_code_raw']).strip()
                        suffix = f" (-{bcode})"
                        if name.endswith(suffix):
                            name = name[:-len(suffix)]
                        return (name, bcode) in sel_inv_tuples
                        
                    export_target_inv = df_inv[df_inv.apply(match_inv_selected, axis=1)]
                    
                    if len(export_target_inv) > 0:
                        export_target_inv = export_target_inv.drop(columns=['branch_code_raw'])
                    else: export_target_inv = None
                    
                    if export_target_inv is not None:
                        st.info(f"✅ **{len(export_target_inv)}** records selected for export.")
                        
                        ci1, ci2, ci3 = st.columns([4, 2, 2])
                        exp_opt3 = ci1.selectbox(
                            "📤 Select Export Format",
                            [
                                "🏢 Internal Export - PDF (Combined)",
                                "🏢 Internal Export - Excel (Combined)",
                                "🤝 Client Export - ZIP (PDFs by Group)",
                                "🤝 Client Export - ZIP (Excels by Group)",
                                "📄 External Export - PDF (No Group)",
                                "📄 External Export - Excel (No Group)"
                            ],
                            label_visibility="collapsed",
                            key="sel_exp3"
                        )
                        
                        if ci2.button("🚀 Generate Report", use_container_width=True, key="btn_gen3"):
                            with st.spinner("Generating... Please wait..."):
                                now_d = datetime.now(HKT).strftime('%Y%m%d')
                                if "Internal Export - PDF" in exp_opt3:
                                    st.session_state.ex_data3 = generate_inv_pdf(export_target_inv, str(target_year), month_str)
                                    st.session_state.ex_name3 = f"Invoicing_Internal_{target_year}_{month_str}_{now_d}.pdf"
                                    st.session_state.ex_mime3 = "application/pdf"
                                elif "Internal Export - Excel" in exp_opt3:
                                    st.session_state.ex_data3 = generate_inv_excel(export_target_inv, str(target_year), month_str)
                                    st.session_state.ex_name3 = f"Invoicing_Internal_{target_year}_{month_str}_{now_d}.xlsx"
                                    st.session_state.ex_mime3 = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                elif "Client Export - ZIP (PDFs" in exp_opt3:
                                    st.session_state.ex_data3 = create_zip_pdfs(export_target_inv, "Invoicing", str(target_year), month_str)
                                    st.session_state.ex_name3 = f"Invoicing_PDF_{target_year}_{month_str}_{now_d}.zip"
                                    st.session_state.ex_mime3 = "application/zip"
                                elif "Client Export - ZIP (Excels" in exp_opt3:
                                    st.session_state.ex_data3 = create_zip_excels(export_target_inv, "Invoicing", str(target_year), month_str)
                                    st.session_state.ex_name3 = f"Invoicing_Excel_{target_year}_{month_str}_{now_d}.zip"
                                    st.session_state.ex_mime3 = "application/zip"
                                elif "External Export - PDF" in exp_opt3:
                                    st.session_state.ex_data3 = generate_inv_pdf(export_target_inv, str(target_year), month_str, hide_client_group=True)
                                    st.session_state.ex_name3 = f"Invoicing_External_{target_year}_{month_str}_{now_d}.pdf"
                                    st.session_state.ex_mime3 = "application/pdf"
                                elif "External Export - Excel" in exp_opt3:
                                    st.session_state.ex_data3 = generate_inv_excel(export_target_inv, str(target_year), month_str, hide_client_group=True)
                                    st.session_state.ex_name3 = f"Invoicing_External_{target_year}_{month_str}_{now_d}.xlsx"
                                    st.session_state.ex_mime3 = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

                        if 'ex_data3' in st.session_state:
                            ci3.download_button(
                                "📥 Click to Download", 
                                data=st.session_state.ex_data3, 
                                file_name=st.session_state.ex_name3, 
                                mime=st.session_state.ex_mime3,
                                type="primary", 
                                use_container_width=True, 
                                key="btn_dl3"
                            )
                    else:
                        st.info("💡 Select checkboxes to export specific companies.")
                else:
                    st.info("No records match the current filter.")
            else:
                st.success("🎉 No companies found for this invoicing period.")
                
    else: st.info("No records found.")

# --- 6. Company Register ---
elif choice == "🏢 Company Register":
    st.title("🏢 Company Records Management")
    mode = st.radio("Mode", ["🆕 Add New", "✏️ Edit Existing", "📋 Copy Existing"], horizontal=True)
    df_all = pd.read_sql("SELECT * FROM companies", engine)
    
    df_all['branch_code'] = df_all['branch_code'].fillna('000').astype(str).replace(['', 'None', 'nan', '<NA>'], '000')
    df_all['biz_name'] = df_all['biz_name'].fillna('').astype(str).replace(['None', 'nan', '<NA>'], '')
    
    groups = pd.read_sql("SELECT group_name FROM client_groups", engine)['group_name'].tolist()
    sorted_groups = sorted([g for g in groups if isinstance(g, str)])
    MIN_DATE = datetime(1900, 1, 1)

    d = {'cg': "", 'en': "", 'ch': "", 'biz_name': "", 'br_ref_date': None, 'ar_ref_date': None, 'cessation_date': None, 'place': "", 'p_oth': "", 'idate': None, 'ci': "", 'is_hk_reg': False, 'hk_idate': None, 'hk_ci': "", 'br': "", 'type': "", 'ra': "", 'ca': "", 'rl': "", 'sl': "", 'cl': "", 'agent': 'Vistra', 'year_end': '12/31', 'billing_mode': 'All-in Package', 'n2e': None, 'n2f': None, 'n2d': False, 'n4e': None, 'n4f': None, 'n4d': False, 'nn6_e': None, 'nn6_f': None, 'nn6_d': False, 'dis': None, 'rem': "", 'comp_rec_dict': {}}
    target_name = None
    
    if mode in ["✏️ Edit Existing", "📋 Copy Existing"] and not df_all.empty:
        head_offices = df_all[df_all['branch_code'] == '000']
        sorted_companies = sorted(head_offices['name_en'].unique().tolist())
        
        target_name = st.selectbox("Select Record (Head Office)", [""] + sorted_companies)
        if target_name != "":
            row = head_offices[head_offices['name_en'] == target_name].iloc[0]
            comp_rec = row.get('compliance_records')
            try: comp_rec = json.loads(comp_rec) if isinstance(comp_rec, str) else {}
            except: comp_rec = {}
            if not isinstance(comp_rec, dict): comp_rec = {}
            
            rem_val = str(row.get('remark', ''))
            if rem_val == 'None': rem_val = ""
            
            raw_co_type = str(row.get('co_type', '')).strip()
            if raw_co_type.lower() in ['none', 'nan', '<na>', 'nat', 'null', '']: 
                raw_co_type = ""
            type_options_ref = ["", "Private Company", "Public Company", "Guarantee", "Individual Business", "Non-Hong Kong Company"]
            if raw_co_type and raw_co_type not in type_options_ref:
                raw_lower = raw_co_type.lower()
                if "private" in raw_lower: raw_co_type = "Private Company"
                elif "public" in raw_lower: raw_co_type = "Public Company"
                elif "guarantee" in raw_lower: raw_co_type = "Guarantee"
                elif "individual" in raw_lower or "sole" in raw_lower: raw_co_type = "Individual Business"
                elif "non" in raw_lower or "hk" in raw_lower: raw_co_type = "Non-Hong Kong Company"
                else:
                    matched = next((opt for opt in type_options_ref if opt.lower() == raw_lower), "")
                    raw_co_type = matched
            
            b_mode = str(row.get('billing_mode', 'All-in Package'))
            if 'All-in' in b_mode: b_mode = 'All-in Package'
            elif 'Itemized' in b_mode: b_mode = 'Itemized'

            d = {'cg': row.get('client_group', ""), 'en': row.get('name_en', ""), 'ch': row.get('name_ch', ""), 
                 'biz_name': row.get('biz_name', ""), 
                 'br_ref_date': row.get('br_ref_date'),
                 'ar_ref_date': row.get('ar_ref_date'),
                 'cessation_date': row.get('cessation_date'),
                 'place': row.get('incorp_place', ""), 'p_oth': row.get('incorp_place_others', ""), 'idate': row.get('incorp_date'), 'ci': row.get('ci_no', ""), 'is_hk_reg': str(row.get('is_hk_registered', "")) == 'True', 'hk_idate': row.get('hk_incorp_date'), 'hk_ci': row.get('hk_ci_no', ""), 'br': row.get('br_no', ""), 'type': raw_co_type, 'ra': row.get('reg_addr', ""), 'ca': row.get('corres_addr', ""), 'rl': row.get('round_loc', ""), 'sl': row.get('sign_loc', ""), 'cl': row.get('seal_loc', ""),
                 'agent': str(row.get('agent', 'Vistra')), 'year_end': str(row.get('year_end', '12/31')), 'billing_mode': b_mode,
                 'n2e': row.get('nd2a_eff_date'), 'n2f': row.get('nd2a_file_date'), 'n2d': str(row.get('nd2a_download', "")) == 'True', 'n4e': row.get('nd4_eff_date'), 'n4f': row.get('nd4_file_date'), 'n4d': str(row.get('nd4_download', "")) == 'True', 'nn6_e': row.get('nn6_eff_date'), 'nn6_f': row.get('nn6_file_date'), 'nn6_d': str(row.get('nn6_download', "")) == 'True', 'dis': row.get('dissolution_date'), 'rem': rem_val, 'comp_rec_dict': comp_rec}
            if mode == "📋 Copy Existing": d['en'], d['ch'], d['biz_name'] = "", "", ""

    st.header("General Information")
    c1, c2 = st.columns(2)
    with c1: st.markdown("⚠️ Company English Name :red[(Required!)]"); name_en = st.text_input("EN", value=d['en'], label_visibility="collapsed")
    with c2: st.markdown("Company Chinese Name"); name_ch = st.text_input("CH", value=d['ch'], label_visibility="collapsed")
    st.markdown("⚠️ Select Client Group :red[(Required!)]")
    client_group = st.selectbox("Group", [""] + sorted_groups, index=(sorted_groups.index(d['cg'])+1 if d['cg'] in sorted_groups else 0), label_visibility="collapsed")
    st.write("---") 
    
    with st.expander("⚙️ Advanced / Special Cases"):
        st.info("💡 Note: This section sets the Main Branch data. To add or manage branches (001, 002...), please use '🏢 Branches Management' below.")
        st.markdown("**Business Name**")
        biz_name = st.text_input("BizName", value=d['biz_name'], label_visibility="collapsed", help="If filled, reports will show Business Name: XXX")
        
        st.markdown("---")
        c_adv2, c_adv3 = st.columns(2)
        with c_adv2: 
            st.markdown("**Custom BR Ref. Date**")
            st.caption("(Fill this if BR expiry month differs from incorporation month)")
            br_ref_date = st.date_input("BR_Ref", value=to_date(d['br_ref_date']), min_value=MIN_DATE, label_visibility="collapsed", format="YYYY/MM/DD")
        with c_adv3: 
            st.markdown("**Custom AR Ref. Date**")
            st.caption("(Fill this if AR calculation month differs from incorporation month. System automatically adds 42 days.)")
            ar_ref_date = st.date_input("AR_Ref", value=to_date(d['ar_ref_date']), min_value=MIN_DATE, label_visibility="collapsed", format="YYYY/MM/DD")

    st.write("---") 
    
    place_options = ["", "HK", "BVI", "Cayman Island", "Others"]
    st.markdown("⚠️ Place of Incorporation :red[(Required!)]")
    inc_place = st.selectbox("Place", place_options, index=(place_options.index(d['place']) if d['place'] in place_options else 0), label_visibility="collapsed")
    
    place_others = ""
    if inc_place == "Others": 
        st.markdown("⚠️ Specify Others :red[(Required!)]"); place_others = st.text_input("Others_Input", value=d['p_oth'], label_visibility="collapsed")
    
    is_bvi = inc_place not in ["HK", ""]
    agent_val, year_end_val, billing_val = "", "12/31", ""
    
    if is_bvi:
        st.write("---")
        st.subheader("🌴 Offshore Company Settings")
        ob1, ob2, ob3 = st.columns(3)
        with ob1:
            agent_opts = ["Vistra", "ICS", "TMF", "Others"]
            agent_val = st.selectbox("Registered Agent", agent_opts, index=(agent_opts.index(d['agent']) if d['agent'] in agent_opts else 0))
        with ob2:
            year_end_val = st.text_input("Financial Year End (MM/DD)", value=d['year_end'], help="e.g. 12/31, 03/31")
        with ob3:
            bill_opts = ["All-in Package", "Itemized"]
            billing_val = st.selectbox("Billing Mode", bill_opts, index=(bill_opts.index(d['billing_mode']) if d['billing_mode'] in bill_opts else 0))

    if inc_place:
        disp_place = "Others" if inc_place == "Others" else inc_place
        c3, c4 = st.columns(2)
        with c3: st.markdown(f"⚠️ {disp_place} Incorp Date :red[(Required!)]"); inc_date = st.date_input("Date", value=to_date(d['idate']), min_value=MIN_DATE, label_visibility="collapsed", format="YYYY/MM/DD")
        with c4: st.markdown(f"⚠️ {disp_place} CI Number :red[(Required!)]"); ci_no = st.text_input("CI", value=d['ci'], label_visibility="collapsed")
    else:
        inc_date = None
        ci_no = ""
    
    is_hk_reg, hk_idate, hk_ci, br_no = False, None, "", ""
    if inc_place == "HK":
        st.markdown("⚠️ HK BR Number (8-digit) :red[(Required!)]")
        br_no = st.text_input("BR", value=d['br'], label_visibility="collapsed", help="只須輸入首 8 位數字。分行號碼將於分行管理中自動結合。")
    elif inc_place in ["BVI", "Cayman Island", "Others"]:
        st.write("---")
        is_hk_reg = st.checkbox("Registered as Non-Hong Kong Company in HK?", value=d['is_hk_reg'])
        if is_hk_reg:
            st.info("📌 Hong Kong Registration Details")
            hk1, hk2 = st.columns(2)
            with hk1: st.markdown("⚠️ HK Incorp Date :red[(Required!)]"); hk_idate = st.date_input("HK_Date", value=to_date(d['hk_idate']), min_value=MIN_DATE, label_visibility="collapsed", format="YYYY/MM/DD")
            with hk2: st.markdown("⚠️ HK CI Number :red[(Required!)]"); hk_ci = st.text_input("HK_CI", value=d['hk_ci'], label_visibility="collapsed")
            st.markdown("⚠️ HK BR Number (8-digit) :red[(Required!)]")
            br_no = st.text_input("BR", value=d['br'], label_visibility="collapsed", help="只須輸入首 8 位數字。分行號碼將於分行管理中自動結合。")

    st.write("---") 
    type_options = ["", "Private Company", "Public Company", "Guarantee", "Individual Business", "Non-Hong Kong Company"]
    st.markdown("⚠️ Company Type :red[(Required!)]"); co_type = st.selectbox("Type", type_options, index=(type_options.index(d['type']) if d['type'] in type_options else 0), label_visibility="collapsed")

# ==================== 📅 Dynamic Annual Obligations ====================
    updated_comp_json = {}
    if inc_place == "HK" or is_hk_reg or is_bvi:
        st.write("---"); st.header("📅 Annual Obligations")
        base = hk_idate if is_hk_reg and not is_bvi else inc_date
        comp_json_load = d.get('comp_rec_dict', {})
        incorp_year = base.year if base else None
        
        if base:
            today_cal = datetime.now(HKT).date()
            prev_br_by = 'Firm'
            prev_afr_fee_by = 'Firm'
            prev_es_fee_by = 'Firm'
            
            # V221: 只顯示「上年、今年、下年」三年 Tabs
            disp_tabs = [current_system_year - 1, current_system_year, current_system_year + 1]
            
            for y in active_years:
                y_str = str(y)
                y_data = comp_json_load.get(y_str, {})
                
                # 計算繼承的 Paid By
                if incorp_year and y < incorp_year:
                    val_br_by = 'N/A'
                    val_afr_fee_by = 'N/A'
                    val_es_fee_by = 'N/A'
                else:
                    raw_br = str(y_data.get('fee_by', y_data.get('br_paid_by', ''))).strip()
                    val_br_by = raw_br if raw_br else (prev_br_by if prev_br_by != 'N/A' else 'Firm')
                    
                    raw_afr = str(y_data.get('afr_fee_by', '')).strip()
                    val_afr_fee_by = raw_afr if raw_afr else (prev_afr_fee_by if prev_afr_fee_by != 'N/A' else 'Firm')
                    
                    raw_es = str(y_data.get('es_fee_by', '')).strip()
                    val_es_fee_by = raw_es if raw_es else (prev_es_fee_by if prev_es_fee_by != 'N/A' else 'Firm')
                
                # 如果年份唔喺 3年 範圍入面，靜靜地儲存舊數據落 Database，唔好出 UI
                if y not in disp_tabs:
                    updated_comp_json[y_str] = y_data
                    prev_br_by = val_br_by
                    prev_afr_fee_by = val_afr_fee_by
                    prev_es_fee_by = val_es_fee_by
                    continue
                
                # 喺範圍內，正式顯示 UI
                with st.expander(f"📌 Year {y} Compliance (FY {y})", expanded=(y == current_system_year)):
                    
                    if incorp_year and y < incorp_year:
                        st.info(f"### ⚪ Year {y}: Not Incorporated Yet")
                    elif incorp_year and y == incorp_year:
                        if is_bvi: st.success(f"### ✅ Year {y}: Exempt from AFR and ES Filing (1st Year)")
                        else: st.success(f"### ✅ AR Deadline ({y}): Exempt (1st Year)")
                        
                    if is_bvi:
                        nxt_br = calc_bvi_fee_deadline(base, y)
                        nxt_ar = calc_afr_deadline(year_end_val, y)
                        nxt_es = calc_es_deadline(base, y)
                    else:
                        if br_ref_date: nxt_br = get_anniv(y, br_ref_date.month, br_ref_date.day)
                        else: nxt_br = get_anniv(y, base.month, base.day)
                        nxt_ar = calc_ar_deadline(base, ar_ref_date, y)
                        nxt_es = None
                    
                    br_days = (nxt_br - today_cal).days if nxt_br else 0
                    ar_days = (nxt_ar - today_cal).days if nxt_ar else 0
                    es_days = (nxt_es - today_cal).days if nxt_es else 0
                    
                    ar_cr_status = str(y_data.get('ar_cr_status', 'Pending'))
                    
                    is_fee_paid = to_date(y_data.get('fee_date', y_data.get('br_date'))) is not None
                    is_ar_filed = to_date(y_data.get('ar_date')) is not None
                    is_es_filed = to_date(y_data.get('es_date')) is not None
                    
                    col_m1, col_m2, col_m3 = st.columns([1, 1, 1]) if is_bvi else st.columns([1, 1, 0.01])
                    
                    fee_lbl = "Annual Fee Deadline" if is_bvi else "BR Fee Deadline"
                    
                    with col_m1:
                        if val_br_by == "Client":
                            if is_fee_paid: st.success(f"### 🟢 {fee_lbl} ({y}):\n**{nxt_br.strftime('%Y/%m/%d')}**\n\n✅ **Client (Recorded)**")
                            else: st.success(f"### 🟢 {fee_lbl} ({y}):\n**{nxt_br.strftime('%Y/%m/%d')}**\n\n✅ **Client (Pending Record)**")
                        elif val_br_by == "N/A": 
                            st.success(f"### 🟢 {fee_lbl} ({y}):\n**{nxt_br.strftime('%Y/%m/%d')}**\n\n✅ **N/A**")
                        else:
                            if is_fee_paid:
                                st.success(f"### 🟢 {fee_lbl} ({y}):\n**{nxt_br.strftime('%Y/%m/%d')}**\n\n✅ **Status: Completed**")
                            elif y > today_cal.year: st.info(f"### 🔵 {fee_lbl} ({y}):\n**{nxt_br.strftime('%Y/%m/%d')}**\n\n🕒 Not yet due")
                            elif br_days < 0: st.error(f"### 🚨 {fee_lbl} ({y}):\n**{nxt_br.strftime('%Y/%m/%d')}**\n\n⚠️ **Overdue by {abs(br_days)} days**")
                            elif br_days <= (90 if is_bvi else 30): st.warning(f"### ⏳ {fee_lbl} ({y}):\n**{nxt_br.strftime('%Y/%m/%d')}**\n\n⏰ **Due in {br_days} days**")
                            else: st.success(f"### 🟢 {fee_lbl} ({y}):\n**{nxt_br.strftime('%Y/%m/%d')}**\n\n✅ **Status: Normal**")
                                
                    with col_m2:
                        ar_lbl = "AFR Deadline" if is_bvi else "AR Deadline"
                        if ar_cr_status == 'Exempt (Dormant)': st.success(f"### ✅ {ar_lbl} ({y}): N/A\n\n✅ **Status: Exempt (Dormant)**")
                        elif ar_cr_status == 'Included in Agent Fee': st.success(f"### ✅ {ar_lbl} ({y}): {nxt_ar.strftime('%Y/%m/%d')}\n\n✅ **Status: Included in Agent Fee**")
                        elif y == incorp_year: st.success(f"### ✅ {ar_lbl} ({y}): Exempt")
                        elif is_bvi and is_ar_filed: st.success(f"### 🟢 {ar_lbl} ({y}):\n**{nxt_ar.strftime('%Y/%m/%d')}**\n\n✅ **Status: Completed**")
                        elif y > today_cal.year: st.info(f"### 🔵 {ar_lbl} ({y}):\n**{nxt_ar.strftime('%Y/%m/%d')}**\n\n🕒 Not yet due")
                        elif not is_bvi and ar_cr_status == 'Processing': st.warning(f"### ⏳ {ar_lbl} ({y}):\n**{nxt_ar.strftime('%Y/%m/%d')}**\n\n⏰ **Status: Processing**")
                        elif not is_bvi and ar_cr_status == 'Returned': st.error(f"### 🚨 {ar_lbl} ({y}):\n**{nxt_ar.strftime('%Y/%m/%d')}**\n\n⚠️ **Status: Returned**")
                        elif not is_bvi and ar_cr_status == 'Completed': st.success(f"### 🟢 {ar_lbl} ({y}):\n**{nxt_ar.strftime('%Y/%m/%d')}**\n\n✅ **Status: Completed**")
                        else:
                            if ar_days < 0: st.error(f"### 🚨 {ar_lbl} ({y}):\n**{nxt_ar.strftime('%Y/%m/%d')}**\n\n⚠️ **Overdue by {abs(ar_days)} days**")
                            elif ar_days <= (90 if is_bvi else 72): st.warning(f"### ⏳ {ar_lbl} ({y}):\n**{nxt_ar.strftime('%Y/%m/%d')}**\n\n⏰ **Due in {ar_days} days**")
                            else: st.success(f"### 🟢 {ar_lbl} ({y}):\n**{nxt_ar.strftime('%Y/%m/%d')}**\n\n✅ **Status: Normal**")
                    
                    if is_bvi:
                        with col_m3:
                            if y == incorp_year: st.success(f"### ✅ ES Deadline ({y}): Exempt")
                            elif is_es_filed: st.success(f"### 🟢 ES Deadline ({y}):\n**{nxt_es.strftime('%Y/%m/%d')}**\n\n✅ **Status: Completed**")
                            elif y > today_cal.year: st.info(f"### 🔵 ES Deadline ({y}):\n**{nxt_es.strftime('%Y/%m/%d')}**\n\n🕒 Not yet due")
                            else:
                                if es_days < 0: st.error(f"### 🚨 ES Deadline ({y}):\n**{nxt_es.strftime('%Y/%m/%d')}**\n\n⚠️ **Overdue by {abs(es_days)} days**")
                                elif es_days <= 90: st.warning(f"### ⏳ ES Deadline ({y}):\n**{nxt_es.strftime('%Y/%m/%d')}**\n\n⏰ **Due in {es_days} days**")
                                else: st.success(f"### 🟢 ES Deadline ({y}):\n**{nxt_es.strftime('%Y/%m/%d')}**\n\n✅ **Status: Normal**")
                
                    pay_opts = ["Firm", "Client", "N/A"]
                    
                    if is_bvi:
                        st.markdown("##### 💰 1. Annual Fee")
                        bvi_r1 = st.columns(3)
                        with bvi_r1[0]:
                            br_by = st.selectbox(f"Annual Fee Paid By ({y})", pay_opts, index=(pay_opts.index(val_br_by) if val_br_by in pay_opts else 0), key=f"br_by_{y}")
                            prev_br_by = br_by
                        with bvi_r1[1]:
                            if br_by == "N/A":
                                st.text_input(f"Annual Fee Paid Date ({y})", value="N/A", disabled=True, key=f"br_dt_dis_{y}")
                                l_br = None
                            else:
                                l_br = st.date_input(f"Annual Fee Paid Date ({y})", value=to_date(y_data.get('fee_date', y_data.get('br_date'))), min_value=MIN_DATE, key=f"br_dt_{y}", format="YYYY/MM/DD")
                        
                        st.markdown("##### 📄 2. Annual Financial Return (AFR)")
                        bvi_r2 = st.columns(3)
                        is_all_in = 'All-in' in billing_val
                        with bvi_r2[0]:
                            if is_all_in:
                                afr_fee_by = br_by
                                st.selectbox(f"AFR Fee Paid By ({y})", [afr_fee_by], index=0, disabled=True, key=f"afr_fee_by_dis_{y}")
                            else:
                                afr_fee_by = st.selectbox(f"AFR Fee Paid By ({y})", pay_opts, index=(pay_opts.index(val_afr_fee_by) if val_afr_fee_by in pay_opts else 0), key=f"afr_fee_by_{y}")
                            prev_afr_fee_by = afr_fee_by
                        with bvi_r2[1]:
                            if is_all_in:
                                if l_br:
                                    st.date_input(f"AFR Fee Paid Date ({y})", value=l_br, disabled=True, key=f"afr_dt_dis_allin_{y}", format="YYYY/MM/DD")
                                    l_afr_fee = l_br
                                else:
                                    st.text_input(f"AFR Fee Paid Date ({y})", value="N/A" if br_by == "N/A" else "", disabled=True, key=f"afr_dt_dis_allin_{y}")
                                    l_afr_fee = None
                            else:
                                if afr_fee_by == "N/A":
                                    st.text_input(f"AFR Fee Paid Date ({y})", value="N/A", disabled=True, key=f"afr_dt_dis_{y}")
                                    l_afr_fee = None
                                else:
                                    l_afr_fee = st.date_input(f"AFR Fee Paid Date ({y})", value=to_date(y_data.get('afr_fee_date')), min_value=MIN_DATE, key=f"afr_fee_dt_{y}", format="YYYY/MM/DD")
                        with bvi_r2[2]:
                            if y == incorp_year:
                                st.text_input(f"AFR Filed Date ({y})", value="N/A (Exempt)", disabled=True, key=f"ar_dt_dis_{y}")
                                l_ar = None
                            else:
                                l_ar = st.date_input(f"AFR Filed Date ({y})", value=to_date(y_data.get('ar_date')), min_value=MIN_DATE, key=f"ar_dt_{y}", format="YYYY/MM/DD")
                        
                        st.markdown("##### 📄 3. Economic Substance (ES Filing)")
                        bvi_r3 = st.columns(3)
                        with bvi_r3[0]:
                            es_fee_by = st.selectbox(f"ES Fee Paid By ({y})", pay_opts, index=(pay_opts.index(val_es_fee_by) if val_es_fee_by in pay_opts else 2), key=f"es_fee_by_{y}")
                            prev_es_fee_by = es_fee_by
                        with bvi_r3[1]:
                            if es_fee_by == "N/A":
                                st.text_input(f"ES Fee Paid Date ({y})", value="N/A", disabled=True, key=f"es_fee_dt_dis_{y}")
                                l_es_fee = None
                            else:
                                l_es_fee = st.date_input(f"ES Fee Paid Date ({y})", value=to_date(y_data.get('es_fee_date')), min_value=MIN_DATE, key=f"es_fee_dt_{y}", format="YYYY/MM/DD")
                        with bvi_r3[2]:
                            if y == incorp_year:
                                st.text_input(f"ES Filed Date ({y})", value="N/A (Exempt)", disabled=True, key=f"es_dt_dis_{y}")
                                l_es = None
                            else:
                                l_es = st.date_input(f"ES Filed Date ({y})", value=to_date(y_data.get('es_date')), min_value=MIN_DATE, key=f"es_dt_{y}", format="YYYY/MM/DD")
                                
                        fin_cr = "Pending"
                        fin_es = "Pending"

                    else:
                        st.markdown("##### 💰 1. Business Registration (BR)")
                        hk_r1 = st.columns(3)
                        with hk_r1[0]:
                            br_by = st.selectbox(f"BR Paid By ({y})", pay_opts, index=(pay_opts.index(val_br_by) if val_br_by in pay_opts else 0), key=f"br_by_{y}")
                            prev_br_by = br_by
                        with hk_r1[1]:
                            if br_by == "N/A":
                                st.text_input(f"BR Paid Date ({y})", value="N/A", disabled=True, key=f"br_dt_dis_{y}")
                                l_br = None
                            else:
                                l_br = st.date_input(f"BR Paid Date ({y})", value=to_date(y_data.get('fee_date', y_data.get('br_date'))), min_value=MIN_DATE, key=f"br_dt_{y}", format="YYYY/MM/DD")
                                
                        st.markdown("##### 📄 2. Annual Return (AR)")
                        hk_r2 = st.columns(3)
                        with hk_r2[0]:
                            if y == incorp_year:
                                st.text_input(f"AR Filed Date ({y})", value="N/A (Exempt)", disabled=True, key=f"ar_dt_dis_{y}")
                                l_ar = None
                            elif ar_cr_status == 'Exempt (Dormant)':
                                st.text_input(f"AR Filed Date ({y})", value="N/A (Dormant)", disabled=True, key=f"ar_dt_dis_{y}")
                                l_ar = None
                            else:
                                l_ar = st.date_input(f"AR Filed Date ({y})", value=to_date(y_data.get('ar_date')), min_value=MIN_DATE, key=f"ar_dt_{y}", format="YYYY/MM/DD")
                        with hk_r2[1]:
                            cr_opts = ["Pending", "Processing", "Returned", "Completed", "Exempt (Dormant)"]
                            if y == incorp_year:
                                st.text_input(f"AR CR Status ({y})", value="N/A", disabled=True, key=f"cr_st_dis_{y}")
                                fin_cr = "Pending"
                            else:
                                cr_idx = cr_opts.index(ar_cr_status) if ar_cr_status in cr_opts else 0
                                fin_cr = st.selectbox(f"AR CR Status ({y})", cr_opts, index=cr_idx, key=f"ar_cr_{y}")
                                
                        afr_fee_by, l_afr_fee, es_fee_by, l_es_fee, l_es, fin_es = 'N/A', None, 'N/A', None, None, "Pending"

                    updated_comp_json[y_str] = {
                        "br_paid_by": br_by,
                        "br_date": l_br.strftime('%Y-%m-%d') if l_br else None,
                        "afr_fee_by": afr_fee_by,
                        "afr_fee_date": l_afr_fee.strftime('%Y-%m-%d') if l_afr_fee else None,
                        "ar_date": l_ar.strftime('%Y-%m-%d') if l_ar else None,
                        "ar_cr_status": fin_cr,
                        "es_fee_by": es_fee_by,
                        "es_fee_date": l_es_fee.strftime('%Y-%m-%d') if l_es_fee else None,
                        "es_date": l_es.strftime('%Y-%m-%d') if l_es else None,
                        "es_status": fin_es
                    }

    n2e, n2f, n2d = d['n2e'], d['n2f'], d['n2d']
    n4e, n4f, n4d = d['n4e'], d['n4f'], d['n4d']
    nn6_e, nn6_f, nn6_d = d['nn6_e'], d['nn6_f'], d['nn6_d']

    if inc_place == "HK":
        st.write("---"); st.header("📝 Compliance Filings (Local Company)")
        st.subheader("📑 Company Secretary Appointment (ND2A)")
        cc1, cc2, cc3, cc4 = st.columns([3, 3, 3, 1])
        with cc1: n2e = st.date_input("Effective Date (Appt)", value=to_date(d['n2e']), min_value=MIN_DATE, key="n2e_v213", format="YYYY/MM/DD")
        with cc2: n2f = st.date_input("Filing Date (ND2A)", value=to_date(d['n2f']), min_value=MIN_DATE, key="n2f_v213", format="YYYY/MM/DD")
        with cc3:
            st.info("Statutory Period: 15 days")
            if n2e: n2_deadline = (n2e + timedelta(days=15)); st.markdown(f"**Deadline: :red[{n2_deadline.strftime('%Y/%m/%d')}]**") 
        with cc4: n2d = st.checkbox("Downloaded", value=d['n2d'], key="n2d_v213")
        
        st.subheader("📑 Company Secretary Resignation (ND4)")
        cc5, cc6, cc7, cc8 = st.columns([3, 3, 3, 1])
        with cc5: n4e = st.date_input("Effective Date (Resign)", value=to_date(d['n4e']), min_value=MIN_DATE, key="n4e_v213", format="YYYY/MM/DD")
        with cc6: n4f = st.date_input("Filing Date (ND4)", value=to_date(d['n4f']), min_value=MIN_DATE, key="n4f_v213", format="YYYY/MM/DD")
        with cc7:
            st.info("Statutory Period: 15 days")
            if n4e: n4_deadline = (n4e + timedelta(days=15)); st.markdown(f"**Deadline: :red[{n4_deadline.strftime('%Y/%m/%d')}]**") 
        with cc8: n4d = st.checkbox("Downloaded", value=d['n4d'], key="n4d_v213")
        
    elif is_hk_reg:
        st.write("---"); st.header("📝 Compliance Filings (Non-HK Company)")
        st.subheader("📑 Secretary & Director Changes (NN6)")
        c_nn1, c_nn2, c_nn3, c_nn4 = st.columns([3, 3, 3, 1])
        with c_nn1: nn6_e = st.date_input("Effective Date", value=to_date(d['nn6_e']), min_value=MIN_DATE, key="nn6_e_v213", format="YYYY/MM/DD")
        with c_nn2: nn6_f = st.date_input("Filing Date (NN6)", value=to_date(d['nn6_f']), min_value=MIN_DATE, key="nn6_f_v213", format="YYYY/MM/DD")
        with c_nn3:
            st.info("Statutory Period: 1 Month")
            if nn6_e:
                nn6_deadline = add_one_month(nn6_e)
                st.markdown(f"**Deadline: :red[{nn6_deadline.strftime('%Y/%m/%d')}]**")
        with c_nn4: nn6_d = st.checkbox("Downloaded", value=d['nn6_d'], key="nn6_d_v213")

    st.write("---"); st.subheader("📍 Address & Contact")
    ca1, ca2 = st.columns(2)
    with ca1: st.markdown("⚠️ Registered Office Address :red[(Required!)]"); reg_addr = st.text_area("Reg", value=d['ra'], label_visibility="collapsed")
    with ca2: st.markdown("⚠️ Correspondence Address :red[(Required!)]"); corres_addr = st.text_area("Corres", value=d['ca'], label_visibility="collapsed")
    st.subheader("📔 Seal Storage")
    l1, l2, l3 = st.columns(3)
    with l1: st.markdown("⚠️ Round Chop Location :red[(Required!)]"); round_l = st.text_input("Round", value=d['rl'], label_visibility="collapsed")
    with l2: st.markdown("⚠️ Signature Chop Location :red[(Required!)]"); sign_l = st.text_input("Sign", value=d['sl'], label_visibility="collapsed")
    with l3: st.markdown("⚠️ Common Seal Location :red[(Required!)]"); common_l = st.text_input("Seal", value=d['cl'], label_visibility="collapsed")
    st.write("---"); st.markdown("Company Dissolution Date"); dis_date = st.date_input("Dissolution", value=to_date(d['dis']), min_value=MIN_DATE, label_visibility="collapsed", format="YYYY/MM/DD")
    
    st.write("---"); st.subheader("📌 Remarks")
    remark_input = st.text_area("Remark", value=d['rem'], help="If filled, this remark will show up on reports and dashboards.")
    
    row_v213 = {'client_group': client_group, 'name_en': name_en, 'name_ch': name_ch, 'biz_name': biz_name, 'branch_code': '000', 'br_ref_date': br_ref_date, 'ar_ref_date': ar_ref_date, 'cessation_date': None, 'incorp_place': inc_place, 'incorp_place_others': place_others, 'incorp_date': inc_date, 'ci_no': ci_no, 'is_hk_registered': is_hk_reg, 'hk_incorp_date': hk_idate, 'hk_ci_no': hk_ci, 'br_no': br_no, 'co_type': co_type, 'reg_addr': reg_addr, 'corres_addr': corres_addr, 'round_loc': round_l, 'sign_loc': sign_l, 'seal_loc': common_l, 'agent': agent_val, 'year_end': year_end_val, 'billing_mode': billing_val, 'nd2a_eff_date': n2e, 'nd2a_file_date': n2f, 'nd2a_download': n2d, 'nd4_eff_date': n4e, 'nd4_file_date': n4f, 'nd4_download': n4d, 'nn6_eff_date': nn6_e, 'nn6_file_date': nn6_f, 'nn6_download': nn6_d, 'dissolution_date': dis_date, 'remark': remark_input, 'compliance_records': json.dumps(updated_comp_json)}
    
    if mode == "✏️ Edit Existing" and target_name:
        st.write("---")
        st.header("🏢 Branches Management")
        st.info("💡 The main branch and sub-branches share core data like registered address and incorp date. Updating the main branch below will automatically sync to all branches.")
        
        branches_df = df_all[(df_all['name_en'] == target_name) & (df_all['branch_code'] != '000')].copy()
        updated_branch_biz = {}
        
        if not branches_df.empty:
            for _, br in branches_df.iterrows():
                b_code = str(br['branch_code']).strip()
                b_biz = str(br.get('biz_name', ''))
                b_br_ref = to_date(br.get('br_ref_date'))
                b_cess = to_date(br.get('cessation_date'))
                
                col1, col2, col3, col4, col5 = st.columns([1.5, 2.5, 2.5, 2, 1.5])
                col1.markdown(f"**Branch: -{b_code}**")
                with col2:
                    new_b_biz = st.text_input(f"Business Name", value=b_biz if (b_biz and b_biz != 'None') else '', key=f"edit_bbiz_{b_code}", label_visibility="collapsed", placeholder="Business Name")
                    updated_branch_biz[b_code] = new_b_biz
                col3.markdown(f"BR Ref Date: **{b_br_ref.strftime('%Y/%m/%d') if b_br_ref else 'Default'}**")
                col4.markdown(f"Cessed: **{b_cess.strftime('%Y/%m/%d') if b_cess else 'N/A'}**")
                with col5:
                    if st.button(f"🗑️ Delete", key=f"del_br_{b_code}"):
                        try:
                            with engine.begin() as conn:
                                safe_name = target_name.replace("'", "''")
                                safe_code = b_code.replace("'", "''")
                                conn.execute(text(f"DELETE FROM companies WHERE name_en = '{safe_name}' AND branch_code = '{safe_code}'"))
                            st.success(f"Branch {b_code} deleted!")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Delete Error: {e}")
                            
        with st.expander("➕ Add New Branch"):
            b_c1, b_c2, b_c3, b_c4 = st.columns(4)
            with b_c1: new_bcode = st.text_input("Branch Code (e.g. 001)")
            with b_c2: new_bbiz = st.text_input("Business Name")
            with b_c3: new_br_ref = st.date_input("BR Ref Date", value=None, min_value=MIN_DATE)
            with b_c4: new_cess = st.date_input("Cessation Date", value=None, min_value=MIN_DATE)

            if st.button("Save New Branch", key="btn_add_new_branch"):
                clean_bcode = new_bcode.strip()
                if not clean_bcode or clean_bcode == '000':
                    st.error("❌ Please enter a valid branch code (e.g. 001)")
                else:
                    new_br_row = row_v213.copy()
                    new_br_row['branch_code'] = clean_bcode
                    new_br_row['biz_name'] = new_bbiz.strip()
                    new_br_row['br_ref_date'] = new_br_ref
                    new_br_row['cessation_date'] = new_cess
                    new_br_row['ar_ref_date'] = None
                    new_br_row['compliance_records'] = "{}"
                    try:
                        pd.DataFrame([new_br_row]).to_sql('companies', engine, if_exists='append', index=False)
                        st.success(f"✅ Branch {clean_bcode} Added successfully!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Save Failed: {e}")

    mandatory_fields = {"Client Group": client_group, "English Name": name_en, "Place": inc_place, "Company Type": co_type, "Registered Address": reg_addr, "Correspondence Address": corres_addr, "Round Chop Location": round_l, "Signature Chop Location": sign_l, "Common Seal Location": common_l}
    
    if inc_place:
        mandatory_fields[f"{inc_place} Incorp Date"] = inc_date
        mandatory_fields[f"{inc_place} CI Number"] = ci_no
        if inc_place == "Others": mandatory_fields["Specify Others"] = place_others
        if inc_place == "HK": mandatory_fields["BR Number (8-digit)"] = br_no
        
    if is_hk_reg:
        mandatory_fields["HK Incorp Date"] = hk_idate
        mandatory_fields["HK CI Number"] = hk_ci
        mandatory_fields["HK BR Number (8-digit)"] = br_no
        
    missing = [k for k, v in mandatory_fields.items() if not v or str(v).strip() == ""]

    st.write("---")
    if mode in ["🆕 Add New", "📋 Copy Existing"]:
        if st.button("💾 Save To Cloud", key="btn_save_v213"):
            if missing: st.error(f"❌ Missing mandatory fields: {', '.join(missing)}")
            else:
                try:
                    pd.DataFrame([row_v213]).to_sql('companies', engine, if_exists='append', index=False)
                    st.success("✅ Success!"); st.rerun()
                except Exception as save_err:
                    st.error(f"❌ Save Failed! Error details: {save_err}")
    else:
        u_col, d_col = st.columns(2)
        with u_col.popover("🆙 Update"):
            if st.button("Confirm Update (Sync Main & Branches)", key="btn_update_v213"):
                if missing: st.error(f"❌ Missing mandatory fields: {', '.join(missing)}")
                else:
                    try:
                        df_backup = df_all.copy() 
                        existing_branches = df_all[(df_all['name_en'] == target_name) & (df_all['branch_code'] != '000')].to_dict('records')
                        
                        df_all = df_all[df_all['name_en'] != target_name]
                        
                        insert_list = [row_v213]
                        for br in existing_branches:
                            b_code = str(br.get('branch_code')).strip()
                            br_updated = row_v213.copy()
                            br_updated['branch_code'] = b_code
                            br_updated['biz_name'] = updated_branch_biz.get(b_code, br.get('biz_name'))
                            br_updated['br_ref_date'] = br.get('br_ref_date')
                            br_updated['cessation_date'] = br.get('cessation_date')
                            br_updated['ar_ref_date'] = None
                            br_updated['compliance_records'] = br.get('compliance_records', '{}')
                            insert_list.append(br_updated)
                            
                        df_all.to_sql('companies', engine, if_exists='replace', index=False)
                        pd.DataFrame(insert_list).to_sql('companies', engine, if_exists='append', index=False)
                        st.success("✅ Updated Company & Synchronized all Branches!"); st.rerun()
                    except Exception as trans_err:
                        df_backup.to_sql('companies', engine, if_exists='replace', index=False)
                        st.error(f"🛑 SQL Error Detected! Rollback completed. Details: {trans_err}")
        with d_col.popover("🚨 DELETE"):
            st.error(f"Delete {target_name} and ALL its branches?"); conf_s = st.text_input("Type DELETE", key="single_del_v213")
            if st.button("Confirm Delete Record", disabled=(conf_s != "DELETE"), key="btn_del_single_v213"):
                df_all = df_all[df_all['name_en'] != target_name]
                df_all.to_sql('companies', engine, if_exists='replace', index=False); st.rerun()

# --- 7. Group Management ---
elif choice == "⚙️ Group Management":
    st.header("⚙️ Group Management")
    new_g = st.text_input("New Group Name", key="new_group_input_v213")
    if st.button("Add Group", key="btn_add_group_v213"): pd.DataFrame([{'group_name': new_g}]).to_sql('client_groups', engine, if_exists='append', index=False); st.rerun()
    st.write("---")
    g_df = pd.read_sql("SELECT * FROM client_groups", engine)
    if not g_df.empty:
        g_df = g_df.sort_values(by=['group_name'], na_position='last')
        target = st.selectbox("Select Group", g_df['group_name'].tolist(), key="select_group_manage_v213")
        c1, c2 = st.columns(2)
        with c1.popover("✏️ Rename Group"):
            ren = st.text_input("New Name:", key="rename_input_v213")
            conf_r = st.text_input("Type RENAME", key="rename_confirm_text_v213")
            if st.button("Confirm Rename", disabled=(conf_r != "RENAME"), key="btn_group_rename_v213"):
                comp_df = pd.read_sql("SELECT * FROM companies", engine)
                comp_df.loc[comp_df['client_group'] == target, 'client_group'] = ren
                comp_df.to_sql('companies', engine, if_exists='replace', index=False)
                g_df.replace({target: ren}).to_sql('client_groups', engine, if_exists='replace', index=False); st.rerun()
        with c2.popover("🗑️ Delete Group"):
            if st.button("Confirm Delete Group", key="btn_group_delete_v213"): 
                g_df[g_df['group_name'] != target].to_sql('client_groups', engine, if_exists='replace', index=False); st.rerun()

# --- 8. Data Exchange ---
elif choice == "📤 Data Exchange":
    st.header("📤 Data Exchange")
    c1, c2 = st.columns(2)
    now_dx = datetime.now(HKT).strftime('%Y%m%d')
    
    base_cols_t = [c for c in TEMPLATE_COLS if c != 'remark']
    dyn_cols_t = []
    for y in report_years:
        dyn_cols_t.extend([f"{y} Fee Paid By", f"{y} Fee Paid Date", f"{y} AR/AFR Fee Paid By", f"{y} AR/AFR Fee Paid Date", f"{y} AR/AFR Filed Date", f"{y} AR/AFR Status", f"{y} ES Fee Paid By", f"{y} ES Fee Paid Date", f"{y} ES Filed Date"])
    ordered_cols = base_cols_t + dyn_cols_t + ['remark']
    
    df_template = pd.DataFrame(columns=ordered_cols)
    df_template.rename(columns=EXCHANGE_COL_MAPPING, inplace=True)
    
    buf_t = io.BytesIO(); df_template.to_excel(buf_t, index=False)
    c1.download_button(label="📥 Template", data=buf_t.getvalue(), file_name=f"Template_{now_dx}.xlsx")
    
    df_db = pd.read_sql("SELECT * FROM companies", engine)
    df_db['branch_code'] = df_db['branch_code'].fillna('000').astype(str).replace(['', 'None', 'nan', '<NA>'], '000')
    df_db['biz_name'] = df_db['biz_name'].fillna('').astype(str).replace(['None', 'nan', '<NA>'], '')
    
    sort_cols = [c for c in ['client_group', 'name_en', 'branch_code', 'incorp_place'] if c in df_db.columns]
    df_db = df_db.sort_values(by=sort_cols, na_position='last')
    
    export_records = df_db.to_dict('records')
    for row in export_records:
        base_date = get_base_date(row)
        if not base_date and row.get('incorp_place') not in ['HK', '']: base_date = to_date(row.get('incorp_date'))
        incorp_year = base_date.year if base_date else None
        
        branch = str(row.get('branch_code', '000')).strip()
        is_branch = branch != '000'
        cess_date = to_date(row.get('cessation_date'))
        
        comp_rec_str = str(row.get('compliance_records', '{}'))
        try: rec_dict = json.loads(comp_rec_str)
        except: rec_dict = {}
        if not isinstance(rec_dict, dict): rec_dict = {}
        
        prev_br_by = 'Firm'
        prev_afr_fee_by = 'Firm'
        prev_es_fee_by = 'Firm'
        
        for y in report_years:
            y_str = str(y)
            y_data = rec_dict.get(y_str, {})
            
            if incorp_year and y < incorp_year:
                row[f'{y} Fee Paid By'] = 'N/A'
                row[f'{y} AR/AFR Status'] = 'N/A'
                row[f'{y} Fee Paid Date'] = ''
                row[f'{y} AR/AFR Fee Paid By'] = 'N/A'
                row[f'{y} AR/AFR Fee Paid Date'] = ''
                row[f'{y} AR/AFR Filed Date'] = ''
                row[f'{y} ES Fee Paid By'] = 'N/A'
                row[f'{y} ES Fee Paid Date'] = ''
                row[f'{y} ES Filed Date'] = ''
                prev_br_by, prev_afr_fee_by, prev_es_fee_by = 'N/A', 'N/A', 'N/A'
                continue
                
            raw_br_by = str(y_data.get('fee_by', y_data.get('br_paid_by', ''))).strip()
            if raw_br_by: br_by = raw_br_by
            else: br_by = prev_br_by if prev_br_by != 'N/A' else 'Firm'
            
            raw_afr_fee_by = str(y_data.get('afr_fee_by', '')).strip()
            if raw_afr_fee_by: afr_fee_by = raw_afr_fee_by
            else: afr_fee_by = prev_afr_fee_by if prev_afr_fee_by != 'N/A' else 'Firm'
            
            raw_es_fee_by = str(y_data.get('es_fee_by', '')).strip()
            if raw_es_fee_by: es_fee_by = raw_es_fee_by
            else: es_fee_by = prev_es_fee_by if prev_es_fee_by != 'N/A' else 'Firm'
            
            if is_branch and cess_date and y >= cess_date.year:
                br_by = "N/A"
                
            prev_br_by = br_by
            prev_afr_fee_by = afr_fee_by
            prev_es_fee_by = es_fee_by
            
            row[f'{y} Fee Paid By'] = br_by
            
            ar_dt = str(y_data.get('ar_date', ''))
            if ar_dt in ['None', 'nan', '<NA>']: ar_dt = ''
            
            ar_cr_status = y_data.get('ar_cr_status', '')
            if not ar_cr_status:
                if ar_dt: ar_cr_status = 'Completed'
                else: ar_cr_status = 'Pending'
                
            if is_branch:
                row[f'{y} AR/AFR Status'] = "N/A (Branch)"
            else:
                row[f'{y} AR/AFR Status'] = ar_cr_status
            
            br_d = str(y_data.get('fee_date', y_data.get('br_date', '')))
            if br_d in ['None', 'nan', '<NA>']: br_d = ''
            else: br_d = br_d.replace('-', '/')
            
            afr_fee_dt = str(y_data.get('afr_fee_date', ''))
            if afr_fee_dt in ['None', 'nan', '<NA>']: afr_fee_dt = ''
            else: afr_fee_dt = afr_fee_dt.replace('-', '/')
            
            es_fee_dt = str(y_data.get('es_fee_date', ''))
            if es_fee_dt in ['None', 'nan', '<NA>']: es_fee_dt = ''
            else: es_fee_dt = es_fee_dt.replace('-', '/')
            
            es_dt = str(y_data.get('es_date', ''))
            if es_dt in ['None', 'nan', '<NA>']: es_dt = ''
            
            if ar_cr_status in ['Exempt (Dormant)', 'Included in Agent Fee']:
                ar_dt = ''
            elif ar_dt and not is_branch: ar_dt = ar_dt.replace('-', '/')
            elif is_branch: ar_dt = ''
            
            if es_dt and not is_branch: es_dt = es_dt.replace('-', '/')
            elif is_branch: es_dt = ''
            
            if is_branch and cess_date and y >= cess_date.year:
                row[f'{y} Fee Paid Date'] = ""
            else:
                row[f'{y} Fee Paid Date'] = br_d
                
            row[f'{y} AR/AFR Fee Paid By'] = afr_fee_by
            row[f'{y} AR/AFR Fee Paid Date'] = afr_fee_dt
            row[f'{y} AR/AFR Filed Date'] = ar_dt
            row[f'{y} ES Fee Paid By'] = es_fee_by
            row[f'{y} ES Fee Paid Date'] = es_fee_dt
            row[f'{y} ES Filed Date'] = es_dt
            
    df_export = pd.DataFrame(export_records)
    df_export = df_export.reindex(columns=ordered_cols)
    
    for col in ["incorp_date", "hk_incorp_date", "br_ref_date", "ar_ref_date", "cessation_date", "nd2a_eff_date", "nd2a_file_date", "nd4_eff_date", "nd4_file_date", "nn6_eff_date", "nn6_file_date", "dissolution_date"]:
        if col in df_export.columns: df_export[col] = pd.to_datetime(df_export[col], errors='coerce').dt.strftime('%Y/%m/%d')
    
    df_export.rename(columns=EXCHANGE_COL_MAPPING, inplace=True)
    
    buf_e = io.BytesIO()
    df_export.to_excel(buf_e, index=False)
    c2.download_button(label="📦 Export All", data=buf_e.getvalue(), file_name=f"Backup_{now_dx}.xlsx", key="btn_export_all_v213")
    
    st.write("---")
    
    up = st.file_uploader("Upload XLSX to Review Changes", type=["xlsx"], key="file_uploader_v213")
    if up:
        try:
            up_df = pd.read_excel(up, engine='openpyxl', keep_default_na=False)
            up_df.rename(columns=REVERSE_EXCHANGE_MAPPING, inplace=True)
            
            existing_df = pd.read_sql("SELECT * FROM companies", engine)
            
            def clean_branch_code(v):
                v_str = str(v).strip()
                if v_str in ['', 'None', 'nan', '<NA>']: return '000'
                if v_str.endswith('.0'): v_str = v_str[:-2]
                if v_str.isdigit(): return v_str.zfill(3)
                return v_str

            def clean_br_no(v):
                v_str = str(v).strip()
                if not v_str or v_str.lower() in ['none', 'nan', '<na>']: return ''
                if v_str.endswith('.0'): v_str = v_str[:-2]
                if v_str.isdigit() and len(v_str) < 8: return v_str.zfill(8)
                return v_str

            def clean_ci_no(v):
                v_str = str(v).strip()
                if not v_str or v_str.lower() in ['none', 'nan', '<na>']: return ''
                if v_str.endswith('.0'): v_str = v_str[:-2]
                if v_str.isdigit() and len(v_str) < 7: return v_str.zfill(7)
                return v_str

            up_df['branch_code'] = up_df['branch_code'].apply(clean_branch_code)
            existing_df['branch_code'] = existing_df['branch_code'].apply(clean_branch_code)
            if 'br_no' in up_df.columns: up_df['br_no'] = up_df['br_no'].apply(clean_br_no)
            if 'br_no' in existing_df.columns: existing_df['br_no'] = existing_df['br_no'].apply(clean_br_no)
            if 'ci_no' in up_df.columns: up_df['ci_no'] = up_df['ci_no'].apply(clean_ci_no)
            if 'ci_no' in existing_df.columns: existing_df['ci_no'] = existing_df['ci_no'].apply(clean_ci_no)
            if 'hk_ci_no' in up_df.columns: up_df['hk_ci_no'] = up_df['hk_ci_no'].apply(clean_ci_no)
            if 'hk_ci_no' in existing_df.columns: existing_df['hk_ci_no'] = existing_df['hk_ci_no'].apply(clean_ci_no)

            for col in ["incorp_date", "hk_incorp_date", "br_ref_date", "ar_ref_date", "cessation_date", "nd2a_eff_date", "nd2a_file_date", "nd4_eff_date", "nd4_file_date", "nn6_eff_date", "nn6_file_date", "dissolution_date"]:
                if col in up_df.columns: up_df[col] = pd.to_datetime(up_df[col], errors='coerce').dt.date
                if col in existing_df.columns: existing_df[col] = pd.to_datetime(existing_df[col], errors='coerce').dt.date

            validation_errors = []
            for idx, row_new in up_df.iterrows():
                excel_row = idx + 2 
                name_en = str(row_new.get('name_en', 'Unknown')).strip()
                place = str(row_new.get('incorp_place', '')).strip()
                is_hk_reg = str(row_new.get('is_hk_registered', 'False')).strip().lower() in ['true', 'yes', 'y', '1']
                
                missing_fields = []
                if not str(row_new.get('client_group', '')).strip(): missing_fields.append("Client Group")
                if not name_en: missing_fields.append("English Name")
                if not place: missing_fields.append("Place of Incorporation")
                
                if missing_fields:
                    validation_errors.append(f"**Row {excel_row} ({name_en})** - Missing/Invalid fields: :red[{', '.join(missing_fields)}]")

            if validation_errors:
                st.error("🛑 **Upload Failed: Please correct the Excel file and re-upload.**")
                for err in validation_errors:
                    st.markdown(f"- {err}")
            else:
                def get_anchor(r):
                    place = str(r.get('incorp_place', '')).strip()
                    name = str(r.get('name_en', '')).strip()
                    branch = clean_branch_code(r.get('branch_code', '000'))
                    return f"NAME_{name}_BRANCH_{branch}_PLACE_{place}"

                up_df['_anchor'] = up_df.apply(get_anchor, axis=1)
                existing_df['_anchor'] = existing_df.apply(get_anchor, axis=1)
                
                diff_list = []
                for _, row_new in up_df.iterrows():
                    anchor_val = row_new['_anchor']
                    en_name = row_new.get('name_en', 'Unknown')
                    branch_code = clean_branch_code(row_new.get('branch_code', '000'))
                    
                    disp_name = en_name if branch_code == '000' else f"{en_name} [Branch: {branch_code}]"
                    old_row = existing_df[existing_df['_anchor'] == anchor_val]
                    
                    if not old_row.empty:
                        old_row = old_row.iloc[0]
                        for col in TEMPLATE_COLS:
                            if col == 'compliance_records': continue
                            old_v = clean_val(old_row.get(col, ""))
                            new_v = clean_val(row_new.get(col, ""))
                            if old_v != new_v:
                                clean_col_name = EXCHANGE_COL_MAPPING.get(col, col)
                                diff_list.append({"Company": disp_name, "Field": clean_col_name, "Old Value": old_v if old_v else "N/A", "New Value": new_v if new_v else "N/A"})
                                
                        comp_rec_str = str(old_row.get('compliance_records', '{}'))
                        try: rec_dict = json.loads(comp_rec_str)
                        except: rec_dict = {}
                        
                        prev_br_by = 'Firm'
                        prev_afr_fee_by = 'Firm'
                        prev_es_fee_by = 'Firm'
                        
                        for y in report_years:
                            y_str = str(y)
                            y_data = rec_dict.get(y_str, {})
                            
                            old_ar_cr = str(y_data.get('ar_cr_status', ''))
                            old_ar_dt = str(y_data.get('ar_date', ''))
                            if old_ar_dt in ['None', 'nan', '<NA>']: old_ar_dt = ''
                            
                            old_afr_fee_by = str(y_data.get('afr_fee_by', ''))
                            old_afr_fee_dt = str(y_data.get('afr_fee_date', ''))
                            if old_afr_fee_dt in ['None', 'nan', '<NA>']: old_afr_fee_dt = ''
                            
                            old_es_fee_by = str(y_data.get('es_fee_by', ''))
                            old_es_fee_dt = str(y_data.get('es_fee_date', ''))
                            if old_es_fee_dt in ['None', 'nan', '<NA>']: old_es_fee_dt = ''
                            
                            old_es_dt = str(y_data.get('es_date', ''))
                            if old_es_dt in ['None', 'nan', '<NA>']: old_es_dt = ''
                            
                            base_dt = get_base_date(old_row)
                            inc_yr = base_dt.year if base_dt else None
                            
                            if inc_yr and y < inc_yr: old_br_by = 'N/A'
                            else: old_br_by = y_data.get('fee_by', y_data.get('br_paid_by', prev_br_by if prev_br_by != 'N/A' else 'Firm'))
                            prev_br_by = old_br_by
                            
                            old_br_dt = str(y_data.get('fee_date', y_data.get('br_date', '')))
                            if old_br_dt in ['None', 'nan', '<NA>']: old_br_dt = ''
                            
                            new_br_by = str(row_new.get(f'{y} Fee Paid By', row_new.get(f'{y} BR Paid By', ''))).strip()
                            if new_br_by == '': new_br_by = 'Firm'
                            
                            new_ar_cr = str(row_new.get(f'{y} AR/AFR Status', row_new.get(f'{y} AR CR Status', ''))).strip()
                            if new_ar_cr not in ["Pending", "Processing", "Returned", "Completed", "Exempt (Dormant)", "Included in Agent Fee"]: new_ar_cr = "Pending"
                            
                            new_afr_fee_by = str(row_new.get(f'{y} AR/AFR Fee Paid By', '')).strip()
                            if new_afr_fee_by == '': new_afr_fee_by = 'N/A'
                            
                            new_es_fee_by = str(row_new.get(f'{y} ES Fee Paid By', '')).strip()
                            if new_es_fee_by == '': new_es_fee_by = 'N/A'
                            
                            br_date_val = row_new.get(f'{y} Fee Paid Date', row_new.get(f'{y} BR Paid Date', row_new.get(f'{y} BR Date')))
                            afr_fee_date_val = row_new.get(f'{y} AR/AFR Fee Paid Date', '')
                            ar_date_val = row_new.get(f'{y} AR/AFR Filed Date', row_new.get(f'{y} AR Filed Date', row_new.get(f'{y} AR Date')))
                            es_fee_date_val = row_new.get(f'{y} ES Fee Paid Date', '')
                            es_date_val = row_new.get(f'{y} ES Filed Date', '')
                            
                            raw_br_dt = to_date(br_date_val)
                            new_br_dt = raw_br_dt.strftime('%Y-%m-%d') if raw_br_dt else ''
                            raw_afr_fee_dt = to_date(afr_fee_date_val)
                            new_afr_fee_dt = raw_afr_fee_dt.strftime('%Y-%m-%d') if raw_afr_fee_dt else ''
                            raw_ar_dt = to_date(ar_date_val)
                            new_ar_dt = raw_ar_dt.strftime('%Y-%m-%d') if raw_ar_dt else ''
                            raw_es_fee_dt = to_date(es_fee_date_val)
                            new_es_fee_dt = raw_es_fee_dt.strftime('%Y-%m-%d') if raw_es_fee_dt else ''
                            raw_es_dt = to_date(es_date_val)
                            new_es_dt = raw_es_dt.strftime('%Y-%m-%d') if raw_es_dt else ''
                            
                            if old_br_by != new_br_by: diff_list.append({"Company": disp_name, "Field": f"{y} Fee Paid By", "Old Value": old_br_by, "New Value": new_br_by})
                            if old_br_dt != new_br_dt: diff_list.append({"Company": disp_name, "Field": f"{y} Fee Paid Date", "Old Value": old_br_dt.replace('-','/') if old_br_dt else '', "New Value": new_br_dt.replace('-','/') if new_br_dt else ''})
                            if old_afr_fee_by != new_afr_fee_by: diff_list.append({"Company": disp_name, "Field": f"{y} AR/AFR Fee Paid By", "Old Value": old_afr_fee_by, "New Value": new_afr_fee_by})
                            if old_afr_fee_dt != new_afr_fee_dt: diff_list.append({"Company": disp_name, "Field": f"{y} AR/AFR Fee Paid Date", "Old Value": old_afr_fee_dt.replace('-','/') if old_afr_fee_dt else '', "New Value": new_afr_fee_dt.replace('-','/') if new_afr_fee_dt else ''})
                            if old_ar_dt != new_ar_dt: diff_list.append({"Company": disp_name, "Field": f"{y} AR/AFR Filed Date", "Old Value": old_ar_dt.replace('-','/') if old_ar_dt else '', "New Value": new_ar_dt.replace('-','/') if new_ar_dt else ''})
                            if old_ar_cr != new_ar_cr: diff_list.append({"Company": disp_name, "Field": f"{y} AR/AFR Status", "Old Value": old_ar_cr, "New Value": new_ar_cr})
                            if old_es_fee_by != new_es_fee_by: diff_list.append({"Company": disp_name, "Field": f"{y} ES Fee Paid By", "Old Value": old_es_fee_by, "New Value": new_es_fee_by})
                            if old_es_fee_dt != new_es_fee_dt: diff_list.append({"Company": disp_name, "Field": f"{y} ES Fee Paid Date", "Old Value": old_es_fee_dt.replace('-','/') if old_es_fee_dt else '', "New Value": new_es_fee_dt.replace('-','/') if new_es_fee_dt else ''})
                            if old_es_dt != new_es_dt: diff_list.append({"Company": disp_name, "Field": f"{y} ES Filed Date", "Old Value": old_es_dt.replace('-','/') if old_es_dt else '', "New Value": new_es_dt.replace('-','/') if new_es_dt else ''})
                            
                    else:
                        diff_list.append({"Company": disp_name, "Field": "NEW RECORD", "Old Value": "N/A", "New Value": "Will be added"})

                if diff_list or True: 
                    st.subheader("🔍 Review Changes")
                    if diff_list: st.table(pd.DataFrame(diff_list))
                    else: st.info("No changes detected in the file. Click Sync to proceed anyway.")
                    
                    if st.button("🚀 Confirm & Apply Changes", key="btn_final_sync_v213"):
                        new_comp_records = []
                        for idx, row_new in up_df.iterrows():
                            base_dt = get_base_date(row_new)
                            inc_yr = base_dt.year if base_dt else None
                            
                            comp_dict = {}
                            prev_br_by = 'Firm'
                            prev_afr_fee_by = 'Firm'
                            prev_es_fee_by = 'Firm'
                            
                            for y in active_years:
                                y_str = str(y)
                                raw_br_by = str(row_new.get(f'{y} Fee Paid By', row_new.get(f'{y} BR Paid By', ''))).strip()
                                
                                if inc_yr and y < inc_yr: br_by = 'N/A'
                                elif inc_yr and y == inc_yr: br_by = raw_br_by if raw_br_by else 'Firm'
                                else: br_by = raw_br_by if raw_br_by else (prev_br_by if prev_br_by != 'N/A' else 'Firm')
                                prev_br_by = br_by
                                
                                new_ar_cr = str(row_new.get(f'{y} AR/AFR Status', row_new.get(f'{y} AR CR Status', ''))).strip()
                                if new_ar_cr not in ["Pending", "Processing", "Returned", "Completed", "Exempt (Dormant)", "Included in Agent Fee"]: new_ar_cr = "Pending"
                                
                                afr_fee_by = str(row_new.get(f'{y} AR/AFR Fee Paid By', '')).strip()
                                if afr_fee_by == '': afr_fee_by = 'N/A'
                                
                                es_fee_by = str(row_new.get(f'{y} ES Fee Paid By', '')).strip()
                                if es_fee_by == '': es_fee_by = 'N/A'
                                
                                br_date_val = row_new.get(f'{y} Fee Paid Date', row_new.get(f'{y} BR Paid Date', row_new.get(f'{y} BR Date')))
                                afr_fee_date_val = row_new.get(f'{y} AR/AFR Fee Paid Date', '')
                                ar_date_val = row_new.get(f'{y} AR/AFR Filed Date', row_new.get(f'{y} AR Filed Date', row_new.get(f'{y} AR Date')))
                                es_fee_date_val = row_new.get(f'{y} ES Fee Paid Date', '')
                                es_date_val = row_new.get(f'{y} ES Filed Date', '')
                                
                                raw_br = to_date(br_date_val)
                                raw_afr_fee = to_date(afr_fee_date_val)
                                raw_ar = to_date(ar_date_val)
                                raw_es_fee = to_date(es_fee_date_val)
                                raw_es = to_date(es_date_val)
                                
                                if inc_yr and y < inc_yr:
                                    br_by, afr_fee_by, es_fee_by = 'N/A', 'N/A', 'N/A'
                                    raw_br, raw_afr_fee, raw_ar, raw_es_fee, raw_es = None, None, None, None, None
                                elif inc_yr and y == inc_yr:
                                    raw_ar, raw_es = None, None
                                
                                if br_by == 'N/A': raw_br = None
                                if afr_fee_by == 'N/A': raw_afr_fee = None
                                if es_fee_by == 'N/A': raw_es_fee = None
                                if new_ar_cr in ['Exempt (Dormant)', 'Included in Agent Fee']: raw_ar = None
                                    
                                comp_dict[y_str] = {
                                    "br_paid_by": br_by,
                                    "br_date": raw_br.strftime('%Y-%m-%d') if raw_br else None,
                                    "afr_fee_by": afr_fee_by,
                                    "afr_fee_date": raw_afr_fee.strftime('%Y-%m-%d') if raw_afr_fee else None,
                                    "ar_date": raw_ar.strftime('%Y-%m-%d') if raw_ar else None,
                                    "ar_cr_status": new_ar_cr,
                                    "es_fee_by": es_fee_by,
                                    "es_fee_date": raw_es_fee.strftime('%Y-%m-%d') if raw_es_fee else None,
                                    "es_date": raw_es.strftime('%Y-%m-%d') if raw_es else None
                                }
                            new_comp_records.append(json.dumps(comp_dict))
                            
                        up_df['compliance_records'] = new_comp_records
                        up_df['branch_code'] = up_df['branch_code'].apply(clean_branch_code)
                        
                        cols_to_drop = []
                        for y in report_years: cols_to_drop.extend([f"{y} Fee Paid By", f"{y} Fee Paid Date", f"{y} AR/AFR Fee Paid By", f"{y} AR/AFR Fee Paid Date", f"{y} AR/AFR Filed Date", f"{y} AR/AFR Status", f"{y} ES Fee Paid By", f"{y} ES Fee Paid Date", f"{y} ES Filed Date", f"{y} BR Paid By", f"{y} BR Paid Date", f"{y} BR Date", f"{y} AR Filed Date", f"{y} AR Date", f"{y} AR CR Status", f"{y} ES Status"])
                        up_df = up_df.drop(columns=cols_to_drop, errors='ignore')
                        combined_df = pd.concat([existing_df, up_df]).drop_duplicates(subset=['_anchor'], keep='last')
                        combined_df = combined_df.drop(columns=['_anchor'], errors='ignore')
                            
                        combined_df.to_sql('companies', engine, if_exists='replace', index=False)
                        st.success("✅ Sync Completed!")
                        st.balloons()
                        st.rerun()
        except Exception as e: st.error(f"Error: {e}")
